
# inventory_app_v2_2_fixed.py
# POS “Almacén Valentina Soto” – IVA incluido (v2.8.0)
# - SIN layout responsivo (layout fijo)
# - Fechas flexibles y visualización dd-mm-aaaa
# - Cierre de precios por decenas (regla del 5)
# - Registro de ENTRADAS de stock con fecha y observación
# - Exportar Movimientos: Entradas/Ventas/Mermas + resumen mensual por producto + totales del período

import os
import sys
import sqlite3
import re
from datetime import date, datetime, timedelta
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Union, Optional

# ---------------------------- RUTAS / CONSTANTES ----------------------------
def BASE_DIR():
    if hasattr(sys, "_MEIPASS"):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE = BASE_DIR()
DB_PATH = os.path.join(BASE, "inventory.db")
DAILY_DIR = os.path.join(BASE, "ReportesExcel")
HIST_DIR = os.path.join(BASE, "ReportesHistoricos")
os.makedirs(DAILY_DIR, exist_ok=True)
os.makedirs(HIST_DIR, exist_ok=True)

SALES_DAILY_DIR = os.path.join(DAILY_DIR, "Ventas")
SALES_HIST_DIR = os.path.join(HIST_DIR, "Ventas")
MOVS_DAILY_DIR = os.path.join(DAILY_DIR, "Movimientos")
MOVS_HIST_DIR = os.path.join(HIST_DIR, "Movimientos")
for d in (SALES_DAILY_DIR, SALES_HIST_DIR, MOVS_DAILY_DIR, MOVS_HIST_DIR):
    os.makedirs(d, exist_ok=True)

STORE_NAME = "Almacén Valentina Soto"
IVA = 0.19
NEAR_EXP_DAYS = 7
TOP_ENTRY_W_BARCODE = 260
TOP_ENTRY_W_QTY = 90
TOP_BTN_W = 120
BIG_TOTAL_FONT_SIZE = 34
CART_ROW_HEIGHT = 13

# ---------------------------- EXCEL helpers ----------------------------
BLUE_TOP = "22304C"  # sobrio
BLUE_TIT = "344767"  # encabezado
HEAD_GRAY = "E7EAF0"
TOTAL_GRAY = "D9D9D9"
GREEN_SEC = "92D050"

def peso_fmt():
    return '"$" #,##0'

def ws_autowidth(ws):
    for col_cells in ws.columns:
        letter = None
        max_len = 0
        for cell in col_cells:
            if hasattr(cell, "column_letter"):
                letter = cell.column_letter
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        if letter:
            ws.column_dimensions[letter].width = max(max_len + 2, 12)

def style_header(ws, title_text):
    last_letter = get_column_letter(ws.max_column if ws.max_column >= 6 else 6)
    ws.merge_cells(f"A1:{last_letter}1")
    ws["A1"] = STORE_NAME
    ws["A1"].font = Font(size=13, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for c in ws["1:1"]:
        c.fill = PatternFill("solid", fgColor=BLUE_TOP)
    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"] = title_text
    ws["A2"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for c in ws["2:2"]:
        c.fill = PatternFill("solid", fgColor=BLUE_TIT)

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=HEAD_GRAY)
        cell.border = border

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = border

def style_section_row(ws, row_index, total_cols=7, color=GREEN_SEC, text_bold=True):
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row_index, column=col)
        cell.fill = PatternFill("solid", fgColor=color)
        if text_bold:
            cell.font = Font(bold=True)

def add_totals_block(ws, total_sum):
    ws.append([])
    ws.append([])
    neto = int(round(total_sum / (1 + IVA)))
    iva = total_sum - neto
    base_row = ws.max_row + 1
    labels = ["TOTAL NETO", "TOTAL IVA", "TOTAL VENTAS DEL DÍA"]
    values = [neto, iva, total_sum]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor=TOTAL_GRAY)
    for i, (lab, val) in enumerate(zip(labels, values)):
        r = base_row + i
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).border = border

        ws.cell(row=r, column=2, value="$").font = Font(bold=True)
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=2).border = border

        ws.cell(row=r, column=3, value=val).number_format = peso_fmt()
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=3).fill = fill
        ws.cell(row=r, column=3).border = border

# ---------------------------- FORMATOS / PRECIOS ----------------------------
def peso_str(n: Union[int, float]) -> str:
    s = f"{int(round(n)):,}".replace(",", ".")
    return f"$ {s}"

def qty_label_unit(n_units: int) -> str:
    u = int(n_units)
    return f"{u} unidad" if u == 1 else f"{u} unidades"

def qty_label_kg(grams: int) -> str:
    if grams < 1000:
        return f"{grams:,}".replace(",", ".") + " gr"
    kg = grams / 1000.0
    return f"{kg:.3f}".rstrip("0").rstrip(".").replace(".", ",") + " kg"

def parse_decimal_kg(txt: str) -> float:
    t = (txt or "").strip().replace(",", ".")
    try:
        v = float(t)
        return 0.0 if v < 0 else v
    except Exception:
        return 0.0

def round_price_to_tens_rule_5(value) -> int:
    """
    Cierre por decenas con regla del 5:
    0-4 -> baja; 5-9 -> sube.
    """
    try:
        p = int(round(float(value)))
    except Exception:
        return 0
    last = p % 10
    if last < 5:
        return p - last
    return p + (10 - last)

# ---------------------------- FECHAS (flexibles) ----------------------------
DATE_FORMATS = [
    "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%d %m %Y",
    "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y %m %d",
    "%m/%d/%Y"
]

def parse_any_date_to_iso(text: str) -> Optional[str]:
    if not text:
        return None
    s = text.strip().replace(".", "-").replace("/", "-").replace(" ", " ")
    s = s.lower().replace(" de ", " ").replace("del ", " ")
    # intentos directos
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(text.strip(), fmt)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    # heurística básica
    tokens = re.split(r"[\-\s]+", s)
    if len(tokens) == 3 and all(t.isdigit() for t in tokens):
        a, b, c = tokens
        if len(a) == 4:
            try:
                dt = datetime(int(a), int(b), int(c))
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return None
        else:
            try:
                a_i, b_i, c_i = int(a), int(b), int(c)
                if a_i > 12:
                    dt = datetime(c_i, b_i, a_i)  # dd-mm-aaaa
                else:
                    if b_i > 12:
                        dt = datetime(c_i, a_i, b_i)  # mm-dd-aaaa
                    else:
                        dt = datetime(c_i, b_i, a_i)  # por defecto dd-mm-aaaa
                return dt.strftime("%Y-%m-%d")
            except Exception:
                return None
    return None

def format_dd_mm_yyyy(iso_text: Optional[str]) -> str:
    if not iso_text:
        return "-"
    try:
        if len(iso_text) == 10 and iso_text[4] == "-":
            return datetime.strptime(iso_text, "%Y-%m-%d").strftime("%d-%m-%Y")
        iso2 = parse_any_date_to_iso(iso_text)
        if iso2:
            return datetime.strptime(iso2, "%Y-%m-%d").strftime("%d-%m-%Y")
    except Exception:
        pass
    return iso_text

# ---------------------------- UI HELPERS (PRO CELDAS) ----------------------------
def style_tree():
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass

    BASE_BG = "#F7F9FC"
    FIELD_BG = "#F7F9FC"
    TEXT_FG = "#111827"
    BORDER = "#D1D7E0"
    HEADER_BG = "#344767"
    HEADER_FG = "#FFFFFF"
    HEADER_ACTIVE = "#3F5476"
    ZEBRA_BG = "#EDF3FF"
    HOVER_BG = "#E6ECF6"
    SELECT_BG = "#2F80ED"
    SELECT_FG = "#FFFFFF"

    style.configure(
        "Pro.Treeview",
        font=("Segoe UI", 10, "bold"),
        rowheight=26,
        background=BASE_BG,
        fieldbackground=FIELD_BG,
        foreground=TEXT_FG,
        bordercolor=BORDER,
        lightcolor=BORDER,
        darkcolor=BORDER,
        borderwidth=1
    )
    style.map(
        "Pro.Treeview",
        background=[("selected", SELECT_BG)],
        foreground=[("selected", SELECT_FG)]
    )
    style.configure(
        "Pro.Treeview.Heading",
        font=("Segoe UI", 10, "bold"),
        background=HEADER_BG,
        foreground=HEADER_FG,
        relief="flat",
        padding=(10, 8)
    )
    style.map("Pro.Treeview.Heading", background=[("active", HEADER_ACTIVE)])

    style.configure("striped", background=ZEBRA_BG)
    style.configure("hover", background=HOVER_BG)

def stripe(tree: ttk.Treeview):
    for i, iid in enumerate(tree.get_children()):
        tags = list(tree.item(iid, "tags") or [])
        if i % 2 == 1 and "striped" not in tags:
            tags.append("striped")
        tree.item(iid, tags=tuple(tags))

def attach_hover(tree: ttk.Treeview):
    state = {"hover_iid": None}
    def on_motion(event):
        try:
            iid = tree.identify_row(event.y)
            if iid and iid != state["hover_iid"]:
                if state["hover_iid"]:
                    try:
                        tags = list(tree.item(state["hover_iid"], "tags") or [])
                        tags = [t for t in tags if t != "hover"]
                        tree.item(state["hover_iid"], tags=tuple(tags))
                    except Exception:
                        pass
                try:
                    tags = list(tree.item(iid, "tags") or [])
                    if "hover" not in tags:
                        tags.append("hover")
                    tree.item(iid, tags=tuple(tags))
                except Exception:
                    pass
                state["hover_iid"] = iid
        except Exception:
            pass

    def on_leave(_):
        try:
            if state["hover_iid"]:
                try:
                    tags = list(tree.item(state["hover_iid"], "tags") or [])
                    tags = [t for t in tags if t != "hover"]
                    tree.item(state["hover_iid"], tags=tuple(tags))
                except Exception:
                    pass
                state["hover_iid"] = None
        except Exception:
            pass

    tree.bind("<Motion>", on_motion)
    tree.bind("<Leave>", on_leave)

def wrap_treeview(parent):
    box = ctk.CTkFrame(
        parent,
        fg_color="#FFFFFF",
        corner_radius=12,
        border_width=1,
        border_color="#D1D7E0"
    )
    inner = ctk.CTkFrame(box, fg_color="#F1F4F8")
    inner.pack(fill="both", expand=True, padx=6, pady=6)
    host = tk.Frame(inner, bg="#F1F4F8")
    host.pack(fill="both", expand=True)
    tree = ttk.Treeview(host, style="Pro.Treeview", show="headings")
    vs = ttk.Scrollbar(host, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vs.set)
    tree.pack(side="left", fill="both", expand=True)
    vs.pack(side="right", fill="y")
    attach_hover(tree)
    return box, tree

# ---------------------------- AUTOCOMPLETE POPUP ----------------------------
class AutoSuggestPopup:
    def __init__(self, master: ctk.CTk, entry: ctk.CTkEntry, on_select):
        self.master = master
        self.entry = entry
        self.on_select = on_select
        self.visible = False
        self.items_data = []
        
        # El popup será una ventana Toplevel con estilo profesional
        self.popup = tk.Toplevel(master)
        self.popup.overrideredirect(True)
        self.popup.withdraw()
        self.popup.configure(bg="#FFFFFF", highlightthickness=0)
        
        # Agregar un pequeño borde/sombra
        outer = tk.Frame(self.popup, bg="#E0E0E0", highlightthickness=0)
        outer.pack(fill="both", expand=True)
        
        # Frame interior con borde
        frm = tk.Frame(outer, bg="#FFFFFF", highlightthickness=1, highlightbackground="#D0D0D0")
        frm.pack(fill="both", expand=True, padx=1, pady=1)
        
        # Crear listbox con scrollbar
        self.listbox = tk.Listbox(
            frm,
            font=("Segoe UI", 13),
            activestyle="none",
            fg="#1A1A1A",
            bg="#FAFAFA",
            height=8,
            bd=0,
            highlightthickness=0,
            selectmode="single"
        )
        self.scroll = tk.Scrollbar(frm, orient="vertical", command=self.listbox.yview, bg="#F5F5F5", activebackground="#2C6FB7")
        self.listbox.configure(yscrollcommand=self.scroll.set)
        self.listbox.pack(side="left", fill="both", expand=True, padx=2, pady=2)
        self.scroll.pack(side="right", fill="y", padx=1, pady=2)
        
        # Configurar colores de selección
        self.listbox.configure(selectforeground="#FFFFFF", selectbackground="#2C6FB7")

        # Bindings
        self.listbox.bind("<Button-1>", self._click)
        self.listbox.bind("<MouseWheel>", self._mousewheel)
        self.listbox.bind("<Button-4>", lambda e: self._scroll_up())
        self.listbox.bind("<Button-5>", lambda e: self._scroll_down())
        self.listbox.bind("<Up>", self._key_up)
        self.listbox.bind("<Down>", self._key_down)
        self.listbox.bind("<Return>", self._enter)

        self.entry.bind("<Down>", self._focus_down)
        self.entry.bind("<Up>", self._focus_up)
        self.entry.bind("<Escape>", lambda e: self.hide())

    def _scroll_up(self):
        try:
            self.listbox.yview_scroll(-1, "units")
        except:
            pass

    def _scroll_down(self):
        try:
            self.listbox.yview_scroll(1, "units")
        except:
            pass

    def _key_up(self, event):
        try:
            idx = self.listbox.curselection()
            if not idx:
                if self.listbox.size() > 0:
                    self.listbox.selection_set(0)
            else:
                if idx[0] > 0:
                    self.listbox.selection_clear(0, "end")
                    self.listbox.selection_set(idx[0] - 1)
                    self.listbox.activate(idx[0] - 1)
            return "break"
        except:
            pass

    def _key_down(self, event):
        try:
            idx = self.listbox.curselection()
            if not idx:
                if self.listbox.size() > 0:
                    self.listbox.selection_set(0)
            else:
                if idx[0] < self.listbox.size() - 1:
                    self.listbox.selection_clear(0, "end")
                    self.listbox.selection_set(idx[0] + 1)
                    self.listbox.activate(idx[0] + 1)
            return "break"
        except:
            pass

    def _mousewheel(self, event):
        try:
            direction = -1 if event.delta > 0 else 1
            self.listbox.yview_scroll(direction, "units")
            return "break"
        except:
            pass

    def _click(self, event):
        try:
            # Solo procesar si el evento viene del listbox, no del scrollbar
            if event.widget != self.listbox:
                return "break"
            index = self.listbox.nearest(event.y)
            if index is not None and index >= 0:
                self.listbox.selection_clear(0, "end")
                self.listbox.selection_set(index)
                self.listbox.activate(index)
                self._select_index(index, add_to_cart=False)
        except:
            pass
        return "break"

    def _enter(self, event):
        try:
            idxs = self.listbox.curselection()
            if idxs:
                self._select_index(idxs[0], add_to_cart=True)
            return "break"
        except:
            pass

    def _select_index(self, index, add_to_cart=False):
        try:
            text = self.listbox.get(index)
            self.on_select(text, add_to_cart)
            self.hide()
        except:
            pass

    def _focus_down(self, event):
        if not self.visible:
            return "break"
        try:
            if self.listbox.size() > 0:
                self.listbox.focus_set()
                self.listbox.selection_clear(0, "end")
                self.listbox.selection_set(0)
                self.listbox.activate(0)
                return "break"
        except:
            pass

    def _focus_up(self, event):
        if not self.visible:
            return "break"
        try:
            if self.listbox.size() > 0:
                last = self.listbox.size() - 1
                self.listbox.focus_set()
                self.listbox.selection_clear(0, "end")
                self.listbox.selection_set(last)
                self.listbox.activate(last)
                return "break"
        except:
            pass

    def show(self, items):
        try:
            self.items_data = items
            self.listbox.delete(0, "end")
            for it in items:
                self.listbox.insert("end", it)
            
            if not items:
                self.hide()
                return

            # Posicionar el popup debajo del entry
            self.master.update_idletasks()
            x = self.entry.winfo_rootx()
            y = self.entry.winfo_rooty() + self.entry.winfo_height()
            w = max(self.entry.winfo_width(), 300)
            h = min(260, 28 * min(len(items), 9) + 4)
            
            self.popup.geometry(f"{w}x{h}+{x}+{y}")
            self.popup.deiconify()
            self.popup.lift()
            self.visible = True
        except:
            pass

    def hide(self):
        try:
            if self.visible:
                self.popup.withdraw()
                self.visible = False
        except:
            pass

# ---------------------------- DIÁLOGOS ----------------------------
class DiscountReasonDialog(ctk.CTkToplevel):
    def __init__(self, master, default_text=""):
        super().__init__(master)
        self.title("Motivo del descuento")
        self.resizable(False, False)
        self.result = None

        width, height = 520, 220
        self.geometry(f"{width}x{height}")
        self.update_idletasks()
        x = master.winfo_rootx() + master.winfo_width() // 2 - width // 2
        y = master.winfo_rooty() + master.winfo_height() // 2 - height // 2
        self.geometry(f"+{x}+{y}")

        frm = ctk.CTkFrame(self, fg_color="#F1F4F8")
        frm.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(frm, text="Describe el motivo del descuento:", text_color="#111827").pack(anchor="w")
        self.ent = ctk.CTkEntry(frm, width=460)
        self.ent.pack(fill="x", pady=(6, 6))
        if default_text:
            self.ent.insert(0, default_text)

        btns = ctk.CTkFrame(frm, fg_color="#F1F4F8")
        btns.pack(fill="x", pady=8)
        ctk.CTkButton(btns, text="Guardar", fg_color="#2C6FB7", command=self._ok).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Cancelar", fg_color="#666666", command=self._cancel).pack(side="left", padx=6)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self._cancel())
        self.grab_set()
        self.focus_force()
        self.ent.focus_set()

    def _ok(self):
        txt = (self.ent.get() or "").strip()
        if not txt:
            messagebox.showwarning("Motivo", "Por favor, ingresa el motivo del descuento.")
            return
        self.result = txt
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


class QuantityDialog(ctk.CTkToplevel):
    def __init__(self, master, title: str, unit_type: str, placeholder_units="1", placeholder_kg="0,500"):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.value = None
        self.unit_type = unit_type or "unit"

        width, height = 420, 240
        self.geometry(f"{width}x{height}")
        self.update_idletasks()
        x = master.winfo_rootx() + master.winfo_width() // 2 - width // 2
        y = master.winfo_rooty() + master.winfo_height() // 2 - height // 2
        self.geometry(f"+{x}+{y}")

        frame = ctk.CTkFrame(self, fg_color="#F1F4F8")
        frame.pack(fill="both", expand=True, padx=18, pady=18)

        ctk.CTkLabel(frame, text="Ingresa cantidad:", font=("Segoe UI", 16, "bold"), text_color="#111827").pack(pady=(0, 6))
        unit_txt = "Unidad actual: Kilos (precio por kg)" if self.unit_type == "kg" else "Unidad actual: Unidades"
        self.entry = ctk.CTkEntry(frame, width=280, height=36)
        self.entry.pack(pady=(6, 2))
        self.entry.insert(0, placeholder_kg if self.unit_type == "kg" else placeholder_units)
        ctk.CTkLabel(frame, text=unit_txt, font=("Segoe UI", 12), text_color="#444").pack(pady=(2, 10))

        btns = ctk.CTkFrame(frame, fg_color="#F1F4F8")
        btns.pack(fill="x", pady=8)
        ctk.CTkButton(btns, text="Aceptar", fg_color="#2C6FB7", command=self._ok).pack(side="left", expand=True, padx=6)
        ctk.CTkButton(btns, text="Cancelar", fg_color="#666666", command=self._cancel).pack(side="left", expand=True, padx=6)

        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self._cancel())
        self.grab_set()
        self.focus_force()
        self.entry.focus_set()

    def _ok(self):
        txt = (self.entry.get() or "").strip()
        if self.unit_type == "kg":
            grams = int(round(parse_decimal_kg(txt) * 1000))
            if grams <= 0:
                return
            self.value = grams
        else:
            try:
                q = int(float(txt))
            except Exception:
                q = 0
            if q <= 0:
                return
            self.value = q
        self.destroy()

    def _cancel(self):
        self.value = None
        self.destroy()


class MermaDialog(ctk.CTkToplevel):
    TYPES = ["Vencimiento", "Regalo", "Daño", "Rotura", "Otros"]
    def __init__(self, master, products_values):
        super().__init__(master)
        self.title("Registrar merma")
        self.resizable(False, False)
        self.result = None
        self.products_values = products_values

        width, height = 560, 320
        self.geometry(f"{width}x{height}")
        self.update_idletasks()
        x = master.winfo_rootx() + master.winfo_width() // 2 - width // 2
        y = master.winfo_rooty() + master.winfo_height() // 2 - height // 2
        self.geometry(f"+{x}+{y}")

        frm = ctk.CTkFrame(self, fg_color="#F1F4F8")
        frm.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(frm, text="Producto:", text_color="#111827").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.ent_prod = ctk.CTkEntry(frm, width=360, font=("Segoe UI", 15), text_color="#111827")
        self.ent_prod.grid(row=0, column=1, columnspan=2, sticky="w", padx=6, pady=6)
        self.ent_prod.bind("<KeyRelease>", self._on_product_typed)
        self.ent_prod.bind("<Return>", lambda e: self._on_product_selected())
        self._suggest_popup = AutoSuggestPopup(self, self.ent_prod, self._on_suggestion_selected)

        ctk.CTkLabel(frm, text="Tipo de merma:", text_color="#111827").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.cb_type = ctk.CTkComboBox(frm, values=self.TYPES, width=220)
        self.cb_type.set(self.TYPES[0])
        self.cb_type.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Cantidad:", text_color="#111827").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        self.ent_qty = ctk.CTkEntry(frm, width=180)
        self.ent_qty.insert(0, "1")
        self.ent_qty.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Observación:", text_color="#111827").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        self.ent_note = ctk.CTkEntry(frm, width=360)
        self.ent_note.grid(row=3, column=1, columnspan=2, sticky="w", padx=6, pady=6)

        btns = ctk.CTkFrame(frm, fg_color="#F1F4F8")
        btns.grid(row=4, column=0, columnspan=3, pady=12)
        ctk.CTkButton(btns, text="Guardar", fg_color="#C91414", command=self._ok).grid(row=0, column=0, padx=6)
        ctk.CTkButton(btns, text="Cancelar", fg_color="#666666", command=self._cancel).grid(row=0, column=1, padx=6)

        self.grab_set()
        self.focus_force()

    def _on_product_typed(self, event=None):
        text = self.ent_prod.get().strip().lower()
        if not text:
            self._suggest_popup.hide()
            return
        filtered = [p for p in self.products_values if text in p.lower()]
        self._suggest_popup.show(filtered)

    def _on_suggestion_selected(self, text, add_to_cart=False):
        self.ent_prod.delete(0, "end")
        self.ent_prod.insert(0, text)

    def _on_product_selected(self):
        pass

    def _ok(self):
        self.result = {
            "product_text": self.ent_prod.get().strip(),
            "type": self.cb_type.get().strip(),
            "qty_text": self.ent_qty.get().strip(),
            "note": (self.ent_note.get() or "").strip()
        }
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()


class EntradaDialog(ctk.CTkToplevel):
    """Registrar ENTRADA de stock con fecha y observación."""
    def __init__(self, master, products_values):
        super().__init__(master)
        self.title("Registrar ENTRADA de stock")
        self.resizable(False, False)
        self.result = None
        self.products_values = products_values

        width, height = 560, 340
        self.geometry(f"{width}x{height}")
        self.update_idletasks()
        x = master.winfo_rootx() + master.winfo_width() // 2 - width // 2
        y = master.winfo_rooty() + master.winfo_height() // 2 - height // 2
        self.geometry(f"+{x}+{y}")

        frm = ctk.CTkFrame(self, fg_color="#F1F4F8")
        frm.pack(fill="both", expand=True, padx=16, pady=16)

        ctk.CTkLabel(frm, text="Producto:", text_color="#111827").grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.ent_prod = ctk.CTkEntry(frm, width=360, font=("Segoe UI", 15), text_color="#111827")
        self.ent_prod.grid(row=0, column=1, columnspan=2, sticky="w", padx=6, pady=6)
        self.ent_prod.bind("<KeyRelease>", self._on_product_typed)
        self.ent_prod.bind("<Return>", lambda e: self._on_product_selected())
        self._suggest_popup = AutoSuggestPopup(self, self.ent_prod, self._on_suggestion_selected)

        ctk.CTkLabel(frm, text="Cantidad:", text_color="#111827").grid(row=1, column=0, sticky="e", padx=6, pady=6)
        self.ent_qty = ctk.CTkEntry(frm, width=180)
        self.ent_qty.insert(0, "1")
        self.ent_qty.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Fecha (flexible):", text_color="#111827").grid(row=2, column=0, sticky="e", padx=6, pady=6)
        self.ent_date = ctk.CTkEntry(frm, width=180)
        self.ent_date.insert(0, date.today().strftime("%d-%m-%Y"))
        self.ent_date.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Observación:", text_color="#111827").grid(row=3, column=0, sticky="e", padx=6, pady=6)
        self.ent_note = ctk.CTkEntry(frm, width=360)
        self.ent_note.grid(row=3, column=1, columnspan=2, sticky="w", padx=6, pady=6)

        btns = ctk.CTkFrame(frm, fg_color="#F1F4F8")
        btns.grid(row=4, column=0, columnspan=3, pady=12)
        ctk.CTkButton(btns, text="Guardar ENTRADA", fg_color="#2C6FB7", command=self._ok).grid(row=0, column=0, padx=6)
        ctk.CTkButton(btns, text="Cancelar", fg_color="#666666", command=self._cancel).grid(row=0, column=1, padx=6)

        self.grab_set()
        self.focus_force()

    def _on_product_typed(self, event=None):
        text = self.ent_prod.get().strip().lower()
        if not text:
            self._suggest_popup.hide()
            return
        filtered = [p for p in self.products_values if text in p.lower()]
        self._suggest_popup.show(filtered)

    def _on_suggestion_selected(self, text, add_to_cart=False):
        self.ent_prod.delete(0, "end")
        self.ent_prod.insert(0, text)

    def _on_product_selected(self):
        pass

    def _ok(self):
        self.result = {
            "product_text": self.ent_prod.get().strip(),
            "qty_text": self.ent_qty.get().strip(),
            "date_text": self.ent_date.get().strip(),
            "note": (self.ent_note.get() or "").strip()
        }
        self.destroy()

    def _cancel(self):
        self.result = None
        self.destroy()

# ---------------------------- APP ----------------------------
class App(ctk.CTk):
    def on_cart_action_click(self, event):
        region = self.tree_cart.identify("region", event.x, event.y)
        if region != "cell" or not self.tree_cart.identify_row(event.y):
            return
        col = self.tree_cart.identify_column(event.x)
        if col != f"#{len(self.tree_cart['columns'])}":  # Última columna es Acción
            return
        row = self.tree_cart.identify_row(event.y)
        if not row:
            return
        x_cell = event.x - self.tree_cart.bbox(row, col)[0]
        cell_width = self.tree_cart.column("Acción", "width")
        if x_cell < cell_width // 2:
            self.modify_cart_qty(int(row), -1)
        else:
            self.modify_cart_qty(int(row), 1)

    def modify_cart_qty(self, idx, delta):
        if 0 <= idx < len(self.cart):
            it = self.cart[idx]
            new_qty = it["qty"] + delta
            if new_qty < 0:
                del self.cart[idx]
            else:
                it["qty"] = new_qty
            self.render_cart()
    STOCK_RED = 10
    STOCK_YELLOW_MIN = 11
    STOCK_YELLOW_MAX = 15
    KG_ALERT_GRAMS = 700
    PAYMENT_METHODS = ["efectivo", "tarjeta"]

    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        ctk.set_widget_scaling(1.0)
        ctk.set_window_scaling(1.0)
        self.title("POS Inventario - IVA incluido (v2.8.0)")
        self.geometry("1280x880")
        self.minsize(1000, 720)
        self.configure(fg_color="#F1F4F8")

        self.cart = []
        self.discount_value = 0
        self.discount_note = None

        self.db_init()
        self.db_migrations()
        style_tree()

        self.sales_wrap = None
        self.sales_split = None
        self.sales_left = None
        self.sales_right = None

        self.build_ui()
        self.cleared_since = None

        self.refresh_products()
        self.refresh_sales_today()
        self.auto_export_yesterday()

        self.day_win = None
        self.detail_win = None
        self.detail_ticket_id = None
        self.day_win_controls = None
        self.day_win_refresh_job = None
        self.day_win_paused = False

    # ---------------------------- DB ----------------------------
    def _conn(self):
        conn = sqlite3.connect(DB_PATH)
        try:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("PRAGMA journal_mode=WAL")
        except Exception:
            pass
        return conn

    def db_init(self):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS products(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE,
                    category TEXT,
                    stock INTEGER,
                    price INTEGER,
                    expiry TEXT,
                    barcode TEXT,
                    unit_type TEXT DEFAULT 'unit',
                    active INTEGER DEFAULT 1
                )
                """
            )
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS tickets(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    payment_method TEXT NOT NULL DEFAULT 'efectivo',
                    total INTEGER NOT NULL,
                    discount INTEGER DEFAULT 0,
                    discount_note TEXT
                )
                """
            )
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS sales(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    ticket_id INTEGER,
                    product_id INTEGER,
                    qty INTEGER,
                    price INTEGER,
                    date TEXT,
                    created_at TEXT,
                    payment_method TEXT DEFAULT 'efectivo',
                    discount INTEGER DEFAULT 0,
                    discount_note TEXT,
                    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE RESTRICT,
                    FOREIGN KEY(ticket_id) REFERENCES tickets(id) ON DELETE CASCADE
                )
                """
            )
            c.execute(
                """
                CREATE TABLE IF NOT EXISTS movimientos(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    product_id INTEGER NOT NULL,
                    qty INTEGER NOT NULL,
                    kind TEXT NOT NULL, -- 'venta' | 'merma' | 'entrada'
                    note TEXT,
                    date TEXT NOT NULL,
                    FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE RESTRICT
                )
                """
            )

    def db_migrations(self):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("CREATE INDEX IF NOT EXISTS idx_products_barcode ON products(barcode)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_products_name ON products(name)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_products_active ON products(active)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_sales_pid ON sales(product_id)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_sales_date ON sales(date)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_sales_ticket ON sales(ticket_id)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_tickets_date ON tickets(date)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_mov_pid ON movimientos(product_id)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_mov_kind ON movimientos(kind)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_mov_date ON movimientos(date)")
            try:
                c.execute(
                    """
                    UPDATE sales
                    SET created_at = CASE
                        WHEN (created_at IS NULL OR created_at='')
                        THEN date || ' 00:00:00'
                        ELSE created_at
                    END
                    """
                )
            except Exception:
                pass

    # ---------------------------- UI ----------------------------
    def build_ui(self):
        tabs = ctk.CTkTabview(self)
        tabs.pack(fill="both", expand=True, padx=10, pady=10)
        self.tabs = tabs
        self.tab_sales = tabs.add("Ventas")
        self.tab_products = tabs.add("Productos")
        self.tab_history = tabs.add("Historial")

        self.ui_sales()
        self.ui_products()
        self.ui_history()

    def ui_sales(self):
        self.sales_wrap = ctk.CTkFrame(self.tab_sales, fg_color="#FFFFFF")
        self.sales_wrap.pack(fill="both", expand=True, padx=8, pady=8)

        # === SECCIÓN SUPERIOR: Entrada de datos ===
        top = ctk.CTkFrame(self.sales_wrap, fg_color="#FFFFFF")
        top.pack(fill="x", padx=6, pady=(6, 12))
        top.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(top, text="Código de barra:", text_color="#111827").grid(row=0, column=0, padx=6, pady=6, sticky="e")
        self.ent_bar = ctk.CTkEntry(top, width=TOP_ENTRY_W_BARCODE)
        self.ent_bar.grid(row=0, column=1, padx=6, pady=6, sticky="w")
        self.ent_bar.bind("<Return>", self.add_by_barcode)

        ctk.CTkLabel(top, text="Producto:", text_color="#111827").grid(row=0, column=2, padx=6, pady=6, sticky="e")
        self.ent_prod = ctk.CTkEntry(top, width=280, font=("Segoe UI", 15), text_color="#111827")
        self.ent_prod.grid(row=0, column=3, padx=6, pady=6, sticky="ew")
        self.ent_prod.bind("<KeyRelease>", self._on_product_typed_autocomplete)
        self.ent_prod.bind("<Return>", lambda e: self.add_to_cart())
        self._suggest_popup = AutoSuggestPopup(self, self.ent_prod, self._on_suggestion_selected)

        ctk.CTkLabel(top, text="Cantidad:", text_color="#111827").grid(row=0, column=4, padx=6, pady=6, sticky="e")
        self.ent_qty = ctk.CTkEntry(top, width=TOP_ENTRY_W_QTY)
        self.ent_qty.insert(0, "1")
        self.ent_qty.grid(row=0, column=5, padx=6, pady=6, sticky="w")

        self.seg_qty_mode = ctk.CTkSegmentedButton(top, values=["Cantidad (kg)", "Monto ($)"], command=self.on_qty_mode_changed)
        self.seg_qty_mode.set("Cantidad (kg)")
        self.seg_qty_mode.grid(row=1, column=4, columnspan=2, padx=6, pady=(0, 6), sticky="ew")

        ctk.CTkButton(top, text="Agregar", fg_color="#2C6FB7", width=TOP_BTN_W, command=self.add_to_cart).grid(row=0, column=6, padx=(10, 6), pady=6, sticky="ew")
        ctk.CTkButton(top, text="Eliminar", fg_color="#C91414", width=TOP_BTN_W, command=self.remove_from_cart).grid(row=0, column=7, padx=6, pady=6, sticky="ew")

        # === SECCIÓN MEDIA: Carrito + Resumen ===
        self.sales_split = ctk.CTkFrame(self.sales_wrap, fg_color="#FFFFFF", height=180)
        self.sales_split.pack(fill="x", expand=False, padx=6, pady=(0, 2))
        self.sales_split.pack_propagate(False)
        self.sales_split.grid_columnconfigure(0, weight=3)
        self.sales_split.grid_columnconfigure(1, weight=2)
        self.sales_split.grid_rowconfigure(0, weight=1)

        # Carrito (izquierda)
        self.sales_left = ctk.CTkFrame(self.sales_split, fg_color="#FFFFFF")
        self.sales_left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        self.sales_left.grid_rowconfigure(0, weight=1)
        
        cols = ("Item", "Nombre", "Cant.", "Precio", "Total")
        cart_box, self.tree_cart = wrap_treeview(self.sales_left)
        cart_box.pack(fill="both", expand=True, padx=6, pady=0)
        self.tree_cart["columns"] = cols

        # Encabezados (ya definidos en azul en style_tree)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Cascadia Mono", 19, "bold"), padding=12)
        style.configure("Treeview", font=("Cascadia Mono", 18, "bold"), rowheight=48, padding=12)
        for c in cols:
            self.tree_cart.heading(c, text=c, anchor="center")
            if c == "Cant.":
                self.tree_cart.column(c, anchor="center", width=140, minwidth=120)
            else:
                self.tree_cart.column(c, anchor="center", width=150, minwidth=100)

        # Ajuste especial para la columna Nombre
        self.tree_cart.column("Nombre", width=280, anchor="w")
        self.tree_cart.column("Item", width=70, anchor="center")

        # Estirar verticalmente dentro del split (sin mover el botón)
        cart_box.pack(fill="both", expand=True, padx=6, pady=0)
        self.tree_cart["columns"] = cols
        for c in cols:
            self.tree_cart.heading(c, text=c, anchor="center")
            self.tree_cart.column(c, width=140, anchor="center")
        self.tree_cart.column("Nombre", width=280, anchor="w")
        self.tree_cart.column("Item", width=70, anchor="center")

        # Bind para acciones en la columna Acción
        self.tree_cart.bind("<Button-1>", self.on_cart_action_click)

        # Resumen (derecha) - Frame principal que ocupa todo el espacio
        self.sales_right = ctk.CTkFrame(self.sales_split, fg_color="#FFFFFF")
        self.sales_right.grid(row=0, column=1, sticky="nsew")
        
        # Total grande
        box = ctk.CTkFrame(self.sales_right, fg_color="#EDF3FF", corner_radius=12, border_width=1, border_color="#D1D7E0")
        box.pack(fill="x", pady=(0, 12))
        self.lbl_total_big = ctk.CTkLabel(box, text="$ 0", font=("Segoe UI", 28, "bold"), text_color="#2F80ED")
        self.lbl_total_big.pack(padx=8, pady=8)

        # Neto, IVA, Total
        panel = ctk.CTkFrame(self.sales_right, fg_color="#FFFFFF")
        panel.pack(fill="x", pady=(0, 12))
        self.ent_neto = self._tot_row(panel, "Neto (incluido):", 0)
        self.ent_iva = self._tot_row(panel, "IVA (19%):", 1)
        self.ent_total = self._tot_row(panel, "Total:", 2)

        # Descuento
        discount_row = ctk.CTkFrame(self.sales_right, fg_color="#FFFFFF")
        discount_row.pack(fill="x", padx=6, pady=(0, 8))
        discount_row.grid_columnconfigure(2, weight=1)
        
        ctk.CTkLabel(discount_row, text="Aplicar descuento:", text_color="#111827").grid(row=0, column=0, padx=6, sticky="e")
        self.ent_discount = ctk.CTkEntry(discount_row, width=80)
        self.ent_discount.insert(0, "0")
        self.ent_discount.grid(row=0, column=1, padx=6, sticky="w")
        ctk.CTkButton(discount_row, text="Agregar", width=80, fg_color="#2C6FB7", command=self.apply_discount).grid(row=0, column=3, padx=6, sticky="e")

        # Método de pago
        payrow = ctk.CTkFrame(self.sales_right, fg_color="#FFFFFF")
        payrow.pack(fill="x", padx=6, pady=(0, 12))
        payrow.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(payrow, text="Método pago:", text_color="#111827").grid(row=0, column=0, padx=6, sticky="e")
        self.pay_method = ctk.CTkComboBox(payrow, values=self.PAYMENT_METHODS, width=200)
        self.pay_method.set(self.PAYMENT_METHODS[0])
        self.pay_method.grid(row=0, column=1, padx=6, sticky="ew")

        # Botón procesar (al final)
        ctk.CTkButton(self.sales_right, text="Procesar venta", fg_color="#12a150", height=40, command=self.process_sale).pack(fill="x", padx=6, pady=(0, 6))

        # === SECCIÓN FINAL: Tabla de ventas del día + Botones ===
        lowbar = ctk.CTkFrame(self.sales_wrap, fg_color="#FFFFFF")
        lowbar.pack(fill="x", padx=6, pady=(0, 8))
        lowbar.grid_columnconfigure(0, weight=1)
        
        ctk.CTkButton(lowbar, text="Vaciar día (vista)", fg_color="#6B7280", command=self.clear_day_view).grid(row=0, column=0, padx=6, pady=6, sticky="e")
        ctk.CTkButton(lowbar, text="Abrir ventas del día (ventana)", fg_color="#6B5B95", command=self.open_day_sales_window).grid(row=0, column=1, padx=6, pady=6, sticky="e")

        box2 = ctk.CTkFrame(self.sales_wrap, fg_color="#FFFFFF")
        box2.pack(fill="both", expand=True, padx=6, pady=(0, 4))
        cols2 = ("Ticket", "Productos", "Total venta", "Fecha", "Hora", "Método", "Detalle")
        day_box, self.tree_day = wrap_treeview(box2)
        day_box.pack(fill="both", expand=True, padx=2, pady=2)
        self.tree_day["columns"] = cols2
        for c in cols2:
            self.tree_day.heading(c, text=c, anchor="center")
            self.tree_day.column(c, width=140, anchor="center")
        self.tree_day.column("Ticket", width=80, anchor="center")
        self.tree_day.column("Productos", width=480, anchor="w")
        self.tree_day.column("Hora", width=100, anchor="center")
        self.tree_day.column("Método", width=120, anchor="center")
        self.tree_day.column("Detalle", width=120, anchor="center")

        self.tree_day.bind("<<TreeviewSelect>>", self.on_select_day)
        self.tree_day.bind("<Double-Button-1>", self.on_tree_day_click)
        self.tree_day.bind("<Return>", self.on_tree_day_click)

        tot = ctk.CTkFrame(self.sales_wrap, fg_color="#FFFFFF")
        tot.pack(fill="x", pady=(0, 4))
        self.lbl_day_neto = ctk.CTkLabel(tot, text="Neto: $ 0", font=("Segoe UI", 12, "bold"))
        self.lbl_day_iva = ctk.CTkLabel(tot, text="IVA: $ 0", font=("Segoe UI", 12, "bold"))
        self.lbl_day_total = ctk.CTkLabel(tot, text="Total: $ 0", font=("Segoe UI", 14, "bold"), text_color="#2F80ED")
        self.lbl_day_neto.pack(side="left", padx=6)
        self.lbl_day_iva.pack(side="left", padx=12)
        self.lbl_day_total.pack(side="left", padx=14)

    def _tot_row(self, parent, label, row):
        ctk.CTkLabel(parent, text=label, text_color="#111827").grid(row=row, column=0, sticky="e", padx=8, pady=6)
        e = ctk.CTkEntry(parent, width=200)
        e.insert(0, "0")
        e.configure(state="disabled")
        e.grid(row=row, column=1, sticky="w", padx=8, pady=6)
        return e

    def ui_products(self):
        frm = ctk.CTkFrame(self.tab_products, fg_color="#FFFFFF")
        frm.pack(fill="both", expand=True, padx=8, pady=8)

        form = ctk.CTkFrame(frm, fg_color="#FFFFFF")
        form.pack(fill="x", pady=6)

        ctk.CTkLabel(form, text="Nombre:", text_color="#111827").grid(row=0, column=0, padx=6, sticky="e")
        self.p_name = ctk.CTkEntry(form, width=260)
        self.p_name.grid(row=0, column=1, padx=6, sticky="w")

        ctk.CTkLabel(form, text="Categoría:", text_color="#111827").grid(row=0, column=2, padx=6, sticky="e")
        self.p_cat = ctk.CTkEntry(form, width=180)
        self.p_cat.grid(row=0, column=3, padx=6, sticky="w")

        ctk.CTkLabel(form, text="Tipo:", text_color="#111827").grid(row=1, column=0, padx=6, sticky="e")
        self.p_type = ctk.CTkComboBox(form, values=["Unidades", "Kilos (precio por kg)"], width=200)
        self.p_type.set("Unidades")
        self.p_type.grid(row=1, column=1, padx=6, sticky="w")

        ctk.CTkLabel(form, text="Stock:", text_color="#111827").grid(row=1, column=2, padx=6, sticky="e")
        self.p_stock = ctk.CTkEntry(form, width=120)
        self.p_stock.grid(row=1, column=3, padx=6, sticky="w")

        ctk.CTkLabel(form, text="Precio:", text_color="#111827").grid(row=2, column=0, padx=6, sticky="e")
        self.p_price = ctk.CTkEntry(form, width=160)
        self.p_price.grid(row=2, column=1, padx=6, sticky="w")
        self.p_price.bind("<FocusOut>", lambda e: self._apply_price_rounding_entry())

        ctk.CTkLabel(form, text="Vence (flexible):", text_color="#111827").grid(row=2, column=2, padx=6, sticky="e")
        self.p_exp = ctk.CTkEntry(form, width=160)
        self.p_exp.grid(row=2, column=3, padx=6, sticky="w")

        ctk.CTkLabel(form, text="Código barra:", text_color="#111827").grid(row=3, column=0, padx=6, sticky="e")
        self.p_bar = ctk.CTkEntry(form, width=180)
        self.p_bar.grid(row=3, column=1, padx=6, sticky="w")

        btns = ctk.CTkFrame(form, fg_color="#FFFFFF")
        btns.grid(row=4, column=0, columnspan=6, pady=8)
        ctk.CTkButton(btns, text="Guardar", fg_color="#2C6FB7", command=self.product_save).grid(row=0, column=0, padx=5)
        ctk.CTkButton(btns, text="Actualizar", fg_color="#12A150", command=self.product_update).grid(row=0, column=1, padx=5)
        ctk.CTkButton(btns, text="Eliminar", fg_color="#C91414", command=self.product_delete).grid(row=0, column=2, padx=5)
        ctk.CTkButton(btns, text="Refrescar", command=self.refresh_products).grid(row=0, column=3, padx=5)
        ctk.CTkButton(btns, text="Registrar MERMA", fg_color="#C91414", command=self.register_merma).grid(row=0, column=4, padx=5)
        ctk.CTkButton(btns, text="Registrar ENTRADA", fg_color="#2C6FB7", command=self.register_entrada).grid(row=0, column=5, padx=5)

        search = ctk.CTkFrame(frm, fg_color="#FFFFFF")
        search.pack(fill="x", pady=4)
        ctk.CTkLabel(search, text="Buscar:", text_color="#111827").pack(side="left", padx=6)
        self.p_filter = ctk.CTkEntry(search, width=360)
        self.p_filter.pack(side="left", padx=6)
        ctk.CTkButton(search, text="Filtrar", command=self.product_filter).pack(side="left", padx=6)

        cols = ("ID", "Nombre", "Categoría", "Stock", "Precio", "Vencimiento", "Barcode")
        prod_box, self.tree_prod = wrap_treeview(frm)
        self.tree_prod["columns"] = cols
        for c in cols:
            self.tree_prod.heading(c, text=c, anchor="center")
            self.tree_prod.column(c, width=140, anchor="center")
        self.tree_prod.column("ID", width=70, anchor="center")
        self.tree_prod.column("Nombre", width=200, anchor="w")
        prod_box.pack(fill="both", expand=True)

        self.tree_prod.tag_configure("low", background="#FFE7E7", foreground="#7A0022")
        self.tree_prod.tag_configure("warn", background="#FFF4DC", foreground="#6B5200")
        self.tree_prod.tag_configure("expired", background="#FFD6D6", foreground="#B30000")
        self.tree_prod.tag_configure("near", background="#FFEED6", foreground="#8A4B00")
        self.tree_prod.bind("<<TreeviewSelect>>", self.on_pick_product)

    def ui_history(self):
        frm = ctk.CTkFrame(self.tab_history, fg_color="#FFFFFF")
        frm.pack(fill="both", expand=True, padx=8, pady=8)

        ctk.CTkLabel(frm, text="Ventas - Desde (flexible):", text_color="#111827").grid(row=0, column=0, padx=6, pady=6, sticky="e")
        self.h_from = ctk.CTkEntry(frm, width=160)
        self.h_from.grid(row=0, column=1, padx=6, sticky="w")

        ctk.CTkLabel(frm, text="Hasta (flexible):", text_color="#111827").grid(row=0, column=2, padx=6, pady=6, sticky="e")
        self.h_to = ctk.CTkEntry(frm, width=160)
        self.h_to.grid(row=0, column=3, padx=6, sticky="w")

        t = date.today().strftime("%d-%m-%Y")
        self.h_from.insert(0, t)
        self.h_to.insert(0, t)

        ctk.CTkButton(frm, text="Hoy", command=self.hist_today).grid(row=0, column=4, padx=6)
        ctk.CTkButton(frm, text="Buscar Ventas", fg_color="#2C6FB7", command=self.hist_search).grid(row=0, column=5, padx=6)
        ctk.CTkButton(frm, text="Exportar Ventas", fg_color="#12A150", command=self.hist_export).grid(row=0, column=6, padx=6)

        cols = ("Ticket", "Productos", "Total venta", "Fecha", "Método")
        hist_box, self.tree_hist = wrap_treeview(frm)
        self.tree_hist["columns"] = cols
        for c in cols:
            self.tree_hist.heading(c, text=c, anchor="center")
            self.tree_hist.column(c, width=160, anchor="center")
        self.tree_hist.column("Ticket", width=70, anchor="center")
        self.tree_hist.column("Productos", width=480, anchor="w")
        hist_box.grid(row=1, column=0, columnspan=7, sticky="nsew", padx=2, pady=2)

        btn_mov = ctk.CTkButton(frm, text="Exportar Excel Movimientos (Entradas/Ventas/Mermas)", fg_color="#6B5B95", command=self.mov_export_using_sales_range)
        btn_mov.grid(row=2, column=0, columnspan=7, padx=6, pady=(8, 4), sticky="ew")

        frm.grid_columnconfigure(0, weight=1)
        frm.grid_rowconfigure(1, weight=1)

    # ---------------------------- Utilidades ----------------------------
    @staticmethod
    def _to_int(v, default=0):
        try:
            return int(float(v))
        except Exception:
            return default

    @staticmethod
    def fmt_dd(s):
        return format_dd_mm_yyyy(s)

    def get_product(self, pid: int):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE id=?", (pid,))
            return c.fetchone()

    def resolve_product(self, text):
        t = (text or "").strip()
        if not t:
            return None
        # Formato sugerencia: "Nombre — $ Precio"
        if " — " in t:
            name_part = t.split(" — ")[0].strip()
            if name_part:
                with self._conn() as conn:
                    c = conn.cursor()
                    c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE lower(name)=? AND active=1", (name_part.lower(),))
                    row = c.fetchone()
                    if row:
                        return row
        # ID numérico
        if t.isdigit():
            pid = int(t)
            with self._conn() as conn:
                c = conn.cursor()
                c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE id=?", (pid,))
                row = c.fetchone()
                if row:
                    return row
        # nombre exacto o difuso
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE lower(name)=? AND active=1", (t.lower(),))
            row = c.fetchone()
            if row:
                return row
            like = f"%{t.lower()}%"
            c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE lower(name) LIKE ? AND active=1 ORDER BY name ASC", (like,))
            return c.fetchone()

    def ticket_products_text(self, ticket_id: int) -> str:
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                SELECT p.name
                FROM sales s
                JOIN products p ON p.id = s.product_id
                WHERE s.ticket_id=?
                ORDER BY s.id ASC
                """,
                (ticket_id,)
            )
            names = [r[0] for r in c.fetchall()]
            return ", ".join(names) if names else "-"

    # ---------------------------- Modo Cantidad/Monto ($) ----------------------------
    def on_qty_mode_changed(self, value=None):
        mode = self.seg_qty_mode.get()
        self.ent_qty.delete(0, "end")
        if mode.startswith("Monto"):
            self.ent_qty.insert(0, "1500")
        else:
            self.ent_qty.insert(0, "0,500")

    def compute_qty_input(self, unit_type: str, price_per_unit: int, user_text: str) -> int:
        if (unit_type or "unit") == "kg":
            mode = self.seg_qty_mode.get() if hasattr(self, "seg_qty_mode") else "Cantidad (kg)"
            if mode.startswith("Monto"):
                digits = re.sub(r"[^\d\-]", "", (user_text or "0"))
                try:
                    amount = int(digits) if digits not in ("", "-") else 0
                except Exception:
                    amount = 0
                if amount <= 0:
                    return 0
                if int(price_per_unit or 0) <= 0:
                    messagebox.showerror("Precio por kg", "El precio del producto por kilo es 0; no se puede calcular gramos.")
                    return 0
                grams = int(round((amount / int(price_per_unit)) * 1000))
                return max(grams, 1)
            else:
                grams = int(round(parse_decimal_kg(user_text) * 1000))
                return max(grams, 1)
        return self._to_int(user_text, 1)

    # ---------------------------- Carrito / Ventas ----------------------------
    def cart_subtotal(self) -> int:
        subtotal = 0
        for it in self.cart:
            if it["unit_type"] == "kg":
                line = (it["qty"] / 1000.0) * int(it["price"])
            else:
                line = int(it["qty"]) * int(it["price"])
            subtotal += int(round(line))
        return subtotal

    def add_by_barcode(self, _=None):
        code = self.ent_bar.get().strip()
        if not code:
            return
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name,price,expiry,unit_type,stock FROM products WHERE barcode=? AND active=1", (code,))
            row = c.fetchone()
            if not row:
                messagebox.showwarning("Código", "No hay producto con ese código.")
                return
            pid, name, price, exp, unit_type, _stock = row
            qty_for_cart = self.compute_qty_input(unit_type, int(price or 0), self.ent_qty.get())
            if qty_for_cart <= 0:
                return
            if not self.check_stock_before_add(pid, qty_for_cart, name, unit_type):
                return
            self.cart_add(pid, name, qty_for_cart, int(price or 0), unit_type)
            self.ent_bar.delete(0, "end")

    def add_to_cart(self):
        res = self.resolve_product(self.ent_prod.get())
        if not res:
            messagebox.showwarning("Validación", "Selecciona un producto válido.")
            return
        pid, name, price, exp, unit_type, _stock = res
        qty = self.compute_qty_input(unit_type, int(price or 0), self.ent_qty.get())
        if qty <= 0:
            return
        if not self.check_stock_before_add(pid, qty, name, unit_type):
            return
        self.cart_add(pid, name, qty, int(price or 0), unit_type)
        # Restablecer cantidad a 1 después de agregar
        if hasattr(self, 'ent_qty') and self.ent_qty:
            self.ent_qty.delete(0, "end")
            self.ent_qty.insert(0, "1")
        self.ent_prod.delete(0, "end")
        self._suggest_popup.hide()

    def check_stock_before_add(self, pid, qty, name, unit_type):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT stock, unit_type, active FROM products WHERE id=?", (pid,))
            r = c.fetchone()
            if not r:
                messagebox.showerror("Producto", "No existe en base.")
                return False
            stock_db, utype, active = int(r[0] or 0), (r[1] or "unit"), int(r[2] or 1)
            if not active:
                messagebox.showerror("Producto", "Producto inactivo.")
                return False
            if unit_type != utype:
                messagebox.showerror("Producto", "Tipo inconsistente.")
                return False
            in_cart = sum(it["qty"] for it in self.cart if it["pid"] == pid)
            disp = stock_db - in_cart
            if qty > disp:
                msg = f"Stock insuficiente para '{name}'. Disponible: "
                msg += qty_label_kg(max(disp, 0)) if unit_type == "kg" else str(max(disp, 0))
                messagebox.showerror("Stock", msg)
                return False
            return True

    def cart_add(self, pid, name, qty, price, unit_type):
        for it in self.cart:
            if it["pid"] == pid and it["price"] == price and it["unit_type"] == unit_type:
                it["qty"] += qty
                break
        else:
            self.cart.append({"pid": pid, "name": name, "qty": qty, "price": price, "unit_type": unit_type})
        self.render_cart()

    def remove_from_cart(self):
        sel = self.tree_cart.selection()
        if not sel:
            messagebox.showwarning("Eliminar", "Selecciona un ítem del carrito.")
            return
        idx = int(sel[0])
        it = self.cart[idx]
        qdlg = QuantityDialog(self, "Quitar cantidad del carrito", it["unit_type"])
        self.wait_window(qdlg)
        val = qdlg.value
        if val is None or val <= 0:
            return
        if val > it["qty"]:
            messagebox.showerror("Cantidad", "Excede la cantidad que deseas modificar.")
            return
        it["qty"] -= val
        if it["qty"] <= 0:
            del self.cart[idx]
        self.render_cart()

    def apply_discount(self):
        t = (self.ent_discount.get() or "").strip().lower()
        if not t:
            self.discount_value = 0
            self.discount_note = None
            self.render_cart()
            return
        try:
            if t.endswith("%"):
                perc = float(t[:-1].replace(",", "."))
                perc = min(max(perc, 0.0), 100.0)
                disc = int(round(self.cart_subtotal() * (perc / 100.0)))
            else:
                digits = re.sub(r"[^\d\-]", "", t)
                disc = int(digits) if digits not in ("", "-") else 0
            if disc < 0:
                disc = 0
            self.discount_value = min(disc, self.cart_subtotal())
            if self.discount_value > 0:
                dlg = DiscountReasonDialog(self)
                self.wait_window(dlg)
                if not dlg.result:
                    self.discount_value = 0
                    self.discount_note = None
                    self.ent_discount.delete(0, "end")
                    self.ent_discount.insert(0, "0")
                else:
                    self.discount_note = dlg.result
            else:
                self.discount_note = None
            self.render_cart()
        except Exception:
            messagebox.showerror("Descuento", "Valor inválido. Usa monto (ej. 1500) o porcentaje (ej. 10%).")

    def render_cart(self):
        self.tree_cart.delete(*self.tree_cart.get_children())
        subtotal = 0
        for i, it in enumerate(self.cart, start=1):
            if it["unit_type"] == "kg":
                line = (it["qty"] / 1000.0) * int(it["price"])
                qty_val = qty_label_kg(it["qty"])
            else:
                line = int(it["qty"]) * int(it["price"])
                qty_val = str(int(it["qty"]))
            subtotal += int(round(line))
            # Usar símbolos unicode estilizados y más espacio
            qty_txt = f"−   {qty_val}   ＋"
            self.tree_cart.insert("", "end", iid=str(i - 1),
                                  values=(i, it["name"], qty_txt, peso_str(it["price"]), peso_str(line)))
        stripe(self.tree_cart)
        self.update_cart_totals()
    def on_cart_action_click(self, event):
        region = self.tree_cart.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self.tree_cart.identify_column(event.x)
        # La columna de cantidad es la tercera (index 2, columna #3)
        if col != "#3":
            return
        row = self.tree_cart.identify_row(event.y)
        if not row:
            return
        bbox = self.tree_cart.bbox(row, col)
        if not bbox:
            return
        x_cell = event.x - bbox[0]
        cell_width = bbox[2]
        # Si clic en la mitad izquierda, resta; si clic en la mitad derecha, suma
        if x_cell < cell_width / 2:
            self.modify_cart_qty(int(row), -1)
        else:
            self.modify_cart_qty(int(row), 1)


    def modify_cart_qty(self, idx, delta):
        if 0 <= idx < len(self.cart):
            it = self.cart[idx]
            # Obtener stock disponible en base de datos
            with self._conn() as conn:
                c = conn.cursor()
                c.execute("SELECT stock FROM products WHERE id=?", (it["pid"],))
                row = c.fetchone()
                stock_db = int(row[0]) if row and row[0] is not None else 0
            # Calcular cantidad ya en carrito (excluyendo este ítem)
            in_cart_otros = sum(x["qty"] for i, x in enumerate(self.cart) if i != idx and x["pid"] == it["pid"])
            max_qty = stock_db - in_cart_otros
            new_qty = it["qty"] + delta
            # Limitar a stock disponible
            if new_qty > max_qty:
                new_qty = max_qty
            if new_qty <= 0:
                del self.cart[idx]
            else:
                it["qty"] = new_qty
            self.render_cart()

    def update_cart_totals(self):
        subtotal = self.cart_subtotal() if hasattr(self, 'cart_subtotal') else 0
        total_after_discount = max(subtotal - int(getattr(self, 'discount_value', 0) or 0), 0)
        IVA = 0.19
        neto = int(round(total_after_discount / (1 + IVA)))
        iva = total_after_discount - neto

        self._set_entry(self.ent_neto, neto)
        self._set_entry(self.ent_iva, iva)
        self._set_entry(self.ent_total, total_after_discount)
        self.lbl_total_big.configure(text=peso_str(total_after_discount))

    @staticmethod
    def _set_entry(e: ctk.CTkEntry, val):
        e.configure(state="normal")
        e.delete(0, "end")
        e.insert(0, str(val))
        e.configure(state="disabled")

    def product_is_expired(self, pid):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT expiry FROM products WHERE id=?", (pid,))
            r = c.fetchone()
            if not r or not r[0]:
                return False
            try:
                d = datetime.strptime(r[0], "%Y-%m-%d").date()
                return d < date.today()
            except Exception:
                return False

    def process_sale(self):
        if not self.cart:
            messagebox.showwarning("Venta", "No hay ítems para vender.")
            return

        expired = [it["name"] for it in self.cart if self.product_is_expired(it["pid"])]
        if expired:
            if not messagebox.askyesno(
                "Producto vencido",
                "Los siguientes productos están VENCIDOS:\n\n" + ", ".join(expired) + "\n\n¿Continuar igualmente?"
            ):
                return

        txt = (self.ent_discount.get() or "").strip().lower()
        if txt and int(self.discount_value or 0) == 0:
            try:
                if txt.endswith("%"):
                    perc = float(txt[:-1].replace(",", "."))
                    perc = min(max(perc, 0.0), 100.0)
                    disc = int(round(self.cart_subtotal() * (perc / 100.0)))
                else:
                    digits = re.sub(r"[^\d\-]", "", txt)
                    disc = int(digits) if digits not in ("", "-") else 0
                if disc < 0:
                    disc = 0
                self.discount_value = min(disc, self.cart_subtotal())
            except Exception:
                self.discount_value = 0

        if int(self.discount_value or 0) > 0 and not self.discount_note:
            dlg = DiscountReasonDialog(self)
            self.wait_window(dlg)
            if not dlg.result:
                messagebox.showwarning("Motivo requerido", "Debes indicar el motivo del descuento.")
                return
            self.discount_note = dlg.result

        today_iso = date.today().strftime("%Y-%m-%d")
        now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        method = (self.pay_method.get() or "efectivo").strip().lower()
        if method not in self.PAYMENT_METHODS:
            method = "efectivo"

        subtotal = self.cart_subtotal()
        total_after = max(subtotal - int(self.discount_value or 0), 0)

        with self._conn() as conn:
            c = conn.cursor()

            # validación de stock
            for it in self.cart:
                c.execute("SELECT stock, active FROM products WHERE id=?", (it["pid"],))
                r = c.fetchone()
                if (not r) or int(r[1] or 1) == 0 or int(r[0]) < int(it["qty"]):
                    messagebox.showerror("Stock", f"Stock insuficiente o producto inactivo para '{it['name']}'.")
                    conn.rollback()
                    return

            # ticket
            c.execute(
                """
                INSERT INTO tickets(date,created_at,payment_method,total,discount,discount_note)
                VALUES (?,?,?,?,?,?)
                """,
                (today_iso, now_ts, method, int(total_after), int(self.discount_value or 0), self.discount_note or "")
            )
            ticket_id = c.lastrowid

            sid_first = None
            first_it = self.cart[0] if self.cart else None

            # ventas + movimientos + stock
            for idx, it in enumerate(self.cart):
                c.execute(
                    """
                    INSERT INTO sales(ticket_id,product_id,qty,price,date,created_at,payment_method,discount,discount_note)
                    VALUES (?,?,?,?,?,?,?,?,?)
                    """,
                    (ticket_id, it["pid"], int(it["qty"]), int(it["price"]), today_iso, now_ts, method, 0, None)
                )
                if idx == 0:
                    sid_first = c.lastrowid

                c.execute("UPDATE products SET stock=stock-? WHERE id=?", (int(it["qty"]), it["pid"]))
                c.execute(
                    "INSERT INTO movimientos(product_id,qty,kind,note,date) VALUES (?,?,?,?,?)",
                    (
                        it["pid"],
                        int(it["qty"]),
                        "venta",
                        (f"descuento: {self.discount_note}" if self.discount_note and idx == 0 else None),
                        today_iso
                    )
                )

            # descuento aplicado SOLO a la primera línea
            if int(self.discount_value or 0) > 0 and first_it and sid_first:
                if (first_it["unit_type"] or "unit") == "kg":
                    line_total = (int(first_it["qty"]) / 1000.0) * int(first_it["price"])
                else:
                    line_total = int(first_it["qty"]) * int(first_it["price"])
                apply_disc = min(int(self.discount_value), int(round(line_total)))
                c.execute("UPDATE sales SET discount=?, discount_note=? WHERE id=?", (apply_disc, self.discount_note or "", sid_first))

        self.cart.clear()
        self.discount_value = 0
        self.discount_note = None
        self.ent_discount.delete(0, "end")
        self.ent_discount.insert(0, "0")
        self.render_cart()
        self.refresh_products()
        self.refresh_sales_today()
        self.refresh_day_window_if_open()
        self.ent_bar.focus_set()
        messagebox.showinfo("Venta", "Venta procesada en un único ticket.")

    # ---------------------------- Ventas del día (pestaña) ----------------------------
    def refresh_sales_today(self):
        self.tree_day.delete(*self.tree_day.get_children())
        today_iso = date.today().strftime("%Y-%m-%d")
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                SELECT id, date, created_at, payment_method, total
                FROM tickets
                WHERE date=? ORDER BY id ASC
                """,
                (today_iso,)
            )
            rows = c.fetchall()
            subtotal = 0
            for tid, d, created_at, pm, total in rows:
                hora_txt = "--"
                if created_at:
                    try:
                        hora_txt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S").strftime("%H:%M")
                    except Exception:
                        hora_txt = created_at.split(" ")[1] if " " in created_at else created_at
                prod_txt = self.ticket_products_text(tid)
                subtotal += int(total or 0)
                self.tree_day.insert("", "end",
                                     values=(tid, prod_txt, peso_str(total or 0), self.fmt_dd(d), hora_txt, pm or "efectivo", "Ver detalle"))
        stripe(self.tree_day)
        neto = int(round(subtotal / (1 + IVA)))
        iva = subtotal - neto
        self.lbl_day_neto.configure(text=f"Neto: {peso_str(neto)}")
        self.lbl_day_iva.configure(text=f"IVA: {peso_str(iva)}")
        self.lbl_day_total.configure(text=f"Total: {peso_str(subtotal)}")

    def on_tree_day_click(self, event):
        sel = self.tree_day.selection()
        if not sel:
            return
        try:
            values = self.tree_day.item(sel[0], "values")
            if values and len(values) >= 1:
                ticket_id = int(values[0])
                self.show_ticket_detail(ticket_id)
        except Exception:
            pass

    # ---------------------------- DETALLE Ticket ----------------------------
    def show_ticket_detail(self, ticket_id: int):
        win = ctk.CTkToplevel(self)
        win.title(f"Detalle Ticket #{ticket_id}")
        win.geometry("760x620")
        try:
            win.transient(self)
            win.lift()
            win.focus_force()
            win.attributes("-topmost", True)
            win.after(200, lambda: win.attributes("-topmost", False))
        except Exception:
            pass

        frame = ctk.CTkFrame(win, fg_color="#FFFFFF")
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        with self._conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT date, created_at, payment_method, total, discount, discount_note FROM tickets WHERE id=?""", (ticket_id,))
            trow = c.fetchone()
            c.execute(
                """
                SELECT s.id, p.name, s.qty, s.price, p.unit_type, s.discount, s.discount_note
                FROM sales s
                JOIN products p ON p.id = s.product_id
                WHERE s.ticket_id = ?
                ORDER BY s.id ASC
                """,
                (ticket_id,)
            )
            srows = c.fetchall()

        header = ctk.CTkFrame(frame, fg_color="#EDF3FF", corner_radius=12, border_width=1, border_color="#D1D7E0")
        header.pack(fill="x", padx=4, pady=(0, 8))

        d, created_at, pm, total_ticket, tdisc, tnote = (None, None, None, None, None, None)
        if trow:
            d, created_at, pm, total_ticket, tdisc, tnote = trow
        hora_txt = created_at.split(" ")[1] if created_at and " " in created_at else (created_at or "--")
        ctk.CTkLabel(header, text=f"Fecha: {self.fmt_dd(d)} Hora: {hora_txt}", text_color="#111827").pack(anchor="w", padx=8, pady=4)
        ctk.CTkLabel(header, text=f"Método: {pm or 'efectivo'} Total: {peso_str(total_ticket or 0)} Descuento ticket: {peso_str(tdisc or 0)}",
                     text_color="#111827").pack(anchor="w", padx=8, pady=2)
        if (tnote or "").strip():
            ctk.CTkLabel(header, text=f"Motivo descuento: {tnote}", text_color="#374151").pack(anchor="w", padx=8, pady=2)

        edit = ctk.CTkFrame(frame, fg_color="#FFFFFF")
        edit.pack(fill="x", padx=4, pady=(0, 8))
        ctk.CTkLabel(edit, text="Editar ticket", font=("Segoe UI", 14, "bold"), text_color="#2F80ED").grid(row=0, column=0, columnspan=4, sticky="w", padx=8, pady=(4, 8))
        ctk.CTkLabel(edit, text="Método de pago:", text_color="#111827").grid(row=1, column=0, padx=8, pady=4, sticky="e")
        cb_method = ctk.CTkComboBox(edit, values=self.PAYMENT_METHODS, width=160)
        cb_method.set((pm or "efectivo") if (pm or "efectivo") in self.PAYMENT_METHODS else "efectivo")
        cb_method.grid(row=1, column=1, padx=6, pady=4, sticky="w")
        ctk.CTkLabel(edit, text="Descuento ticket:", text_color="#111827").grid(row=1, column=2, padx=8, pady=4, sticky="e")
        ent_disc = ctk.CTkEntry(edit, width=120)
        ent_disc.insert(0, str(int(tdisc or 0)))
        ent_disc.grid(row=1, column=3, padx=6, pady=4, sticky="w")
        ctk.CTkLabel(edit, text="Motivo:", text_color="#111827").grid(row=2, column=0, padx=8, pady=4, sticky="e")
        ent_note = ctk.CTkEntry(edit, width=420)
        ent_note.insert(0, (tnote or ""))
        ent_note.grid(row=2, column=1, columnspan=3, padx=6, pady=4, sticky="w")
        ctk.CTkButton(edit, text="Guardar cambios", fg_color="#12a150",
                      command=lambda: self._save_ticket_edits(ticket_id, cb_method.get(), ent_disc.get(), ent_note.get(), win)).grid(row=3, column=0, columnspan=4, padx=8, pady=(8, 6), sticky="ew")

        cols = ("Producto", "Cantidad", "Precio", "Subtotal línea", "Descuento")
        det_box, tree = wrap_treeview(frame)
        det_box.pack(fill="both", expand=True)
        tree["columns"] = cols
        for ccol in cols:
            tree.heading(ccol, text=ccol, anchor="center")
            tree.column(ccol, width=140, anchor="center")
        tree.column("Producto", width=260, anchor="w")

        total_calc = 0
        for _, name, qty, price, utype, sdisc, sdn in srows:
            if (utype or "unit") == "kg":
                line = (int(qty) / 1000.0) * int(price)
                qty_txt = f"{qty_label_kg(int(qty))}"
            else:
                line = int(qty) * int(price)
                qty_txt = f"{qty_label_unit(int(qty))}"
            line_after = max(int(round(line)) - int(sdisc or 0), 0)
            total_calc += line_after
            tree.insert("", "end", values=(name, qty_txt, peso_str(price), peso_str(line_after), peso_str(sdisc or 0)))
        stripe(tree)

        neto = int(round(total_calc / (1 + IVA)))
        iva = total_calc - neto
        tot = ctk.CTkFrame(frame, fg_color="#FFFFFF")
        tot.pack(fill="x", padx=4, pady=8)
        ctk.CTkLabel(tot, text=f"Neto: {peso_str(neto)}", font=("Segoe UI", 12, "bold")).pack(side="left", padx=6)
        ctk.CTkLabel(tot, text=f"IVA: {peso_str(iva)}", font=("Segoe UI", 12, "bold")).pack(side="left", padx=12)
        ctk.CTkLabel(tot, text=f"Total: {peso_str(total_calc)}", font=("Segoe UI", 14, "bold"), text_color="#2F80ED").pack(side="left", padx=14)

    def _save_ticket_edits(self, ticket_id: int, method_text: str, discount_text: str, note_text: str, win=None):
        disc_input = (discount_text or "").strip().lower()
        subtotal = self._calc_ticket_subtotal_raw(ticket_id)
        if disc_input.endswith("%"):
            try:
                perc = float(disc_input[:-1].replace(",", "."))
                perc = min(max(perc, 0.0), 100.0)
                new_disc = int(round(subtotal * (perc / 100.0)))
            except Exception:
                messagebox.showerror("Descuento", "Porcentaje inválido.")
                return
        else:
            digits = re.sub(r"[^\d\-]", "", disc_input or "0")
            try:
                new_disc = int(digits) if digits not in ("", "-") else 0
            except Exception:
                new_disc = 0
        if new_disc < 0:
            new_disc = 0

        method = (method_text or "efectivo").strip().lower()
        if method not in self.PAYMENT_METHODS:
            method = "efectivo"

        note = (note_text or "").strip()
        if new_disc > 0 and not note:
            dlg = DiscountReasonDialog(self)
            self.wait_window(dlg)
            if not dlg.result:
                messagebox.showwarning("Motivo requerido", "Debes indicar el motivo del descuento.")
                return
            note = dlg.result

        ok = self.update_ticket(ticket_id, method, new_disc, note)
        if not ok:
            return
        self.refresh_sales_today()
        self.refresh_day_window_if_open()
        if win and win.winfo_exists():
            try:
                win.destroy()
            except Exception:
                pass
        self.show_ticket_detail(ticket_id)
        messagebox.showinfo("Ticket", "Cambios guardados correctamente.")

    def _calc_ticket_subtotal_raw(self, ticket_id: int) -> int:
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                SELECT s.qty, s.price, p.unit_type
                FROM sales s
                JOIN products p ON p.id = s.product_id
                WHERE s.ticket_id=?
                """,
                (ticket_id,)
            )
            rows = c.fetchall()
        subtotal = 0
        for qty, price, utype in rows:
            if (utype or "unit") == "kg":
                line = (int(qty) / 1000.0) * int(price)
            else:
                line = int(qty) * int(price)
            subtotal += int(round(line))
        return subtotal

    def update_ticket(self, ticket_id: int, new_method: str, new_discount: int, new_note: str) -> bool:
        try:
            with self._conn() as conn:
                c = conn.cursor()
                c.execute(
                    """
                    SELECT s.id, s.qty, s.price, p.unit_type
                    FROM sales s
                    JOIN products p ON p.id = s.product_id
                    WHERE s.ticket_id=?
                    ORDER BY s.id ASC
                    """,
                    (ticket_id,)
                )
                rows = c.fetchall()
                if not rows:
                    messagebox.showerror("Ticket", "No existen ventas en el ticket.")
                    return False

                subtotal = 0
                first_sid, first_qty, first_price, first_utype = rows[0]
                for sid, qty, price, utype in rows:
                    if (utype or "unit") == "kg":
                        line = (int(qty) / 1000.0) * int(price)
                    else:
                        line = int(qty) * int(price)
                    subtotal += int(round(line))

                if (first_utype or "unit") == "kg":
                    first_line_total = (int(first_qty) / 1000.0) * int(first_price)
                else:
                    first_line_total = int(first_qty) * int(first_price)
                apply_disc = min(int(new_discount or 0), int(round(first_line_total)))

            with self._conn() as conn2:
                c2 = conn2.cursor()
                c2.execute("UPDATE sales SET discount=0, discount_note=NULL WHERE ticket_id=?", (ticket_id,))
                if apply_disc > 0:
                    c2.execute("UPDATE sales SET discount=?, discount_note=? WHERE id=?", (apply_disc, (new_note or ""), first_sid))

                new_method_final = (new_method or "efectivo").strip().lower()
                if new_method_final not in self.PAYMENT_METHODS:
                    new_method_final = "efectivo"
                c2.execute("UPDATE sales SET payment_method=? WHERE ticket_id=?", (new_method_final, ticket_id))

                new_total = max(int(subtotal) - int(new_discount or 0), 0)
                c2.execute(
                    """
                    UPDATE tickets
                    SET payment_method=?, discount=?, discount_note=?, total=?
                    WHERE id=?
                    """,
                    (new_method_final, int(new_discount or 0), (new_note or ""), int(new_total), ticket_id)
                )
                conn2.commit()
                return True
        except Exception as e:
            messagebox.showerror("Ticket", f"No se pudo actualizar el ticket:\n{e}")
            return False

    def on_select_day(self, _=None):
        sel = self.tree_day.selection()
        if not sel:
            return
        self.ed_price.delete(0, "end")
        self.ed_qty.delete(0, "end")

    def update_day_sale(self):
        messagebox.showinfo(
            "Edición no disponible",
            "Ahora las ventas se agrupan en un único ticket.\n"
            "La edición por ticket no está habilitada en esta vista.\n"
            "Si la necesitas, te la implemento."
        )

    def delete_day_sale(self):
        sel = self.tree_day.selection()
        if not sel:
            messagebox.showwarning("Eliminar", "Selecciona una venta (ticket).")
            return
        if not messagebox.askyesno("Eliminar", "¿Eliminar este TICKET y restaurar stock de todos sus productos?"):
            return
        tid = int(self.tree_day.item(sel[0], "values")[0])
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT product_id, qty FROM sales WHERE ticket_id=?", (tid,))
            for pid, q in c.fetchall():
                c.execute("UPDATE products SET stock=stock+? WHERE id=?", (int(q), int(pid)))
            c.execute("DELETE FROM sales WHERE ticket_id=?", (tid,))
            c.execute("DELETE FROM tickets WHERE id=?", (tid,))
            conn.commit()
        self.refresh_products()
        self.refresh_sales_today()

    def clear_day_view(self):
        if not messagebox.askyesno("Vaciar vista", "¿Ocultar las ventas anteriores desde este momento?\nLas nuevas ventas se mostrarán normalmente."):
            return
        self.cleared_since = datetime.now()
        self.refresh_sales_today()

    # ---------------------------- MERMAS ----------------------------
    def register_merma(self):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name FROM products WHERE active=1 ORDER BY id DESC")
            rows = c.fetchall()
            values = [f"{name}" for pid, name in rows]
        dlg = MermaDialog(self, values)
        self.wait_window(dlg)
        if not dlg.result:
            return
        res = self.resolve_product(dlg.result["product_text"])
        if not res:
            messagebox.showerror("Merma", "Producto inválido.")
            return
        pid, name, price, exp, unit_type, stock_db = res
        qty = int(round(parse_decimal_kg(dlg.result["qty_text"]) * 1000)) if (unit_type or "unit") == "kg" \
            else self._to_int(dlg.result["qty_text"], 0)
        if qty <= 0:
            messagebox.showerror("Merma", "Cantidad inválida.")
            return
        if qty > int(stock_db or 0):
            messagebox.showerror("Merma", "Stock insuficiente.")
            return
        today_iso = date.today().strftime("%Y-%m-%d")
        tipo = dlg.result["type"] or "Otros"
        nota = dlg.result["note"] or None
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("UPDATE products SET stock=stock-? WHERE id=?", (qty, pid))
            c.execute("INSERT INTO movimientos(product_id,qty,kind,note,date) VALUES (?,?,?,?,?)",
                      (pid, qty, "merma", (tipo + (f" - {nota}" if nota else "")), today_iso))
        self.refresh_products()
        messagebox.showinfo("Merma", "Merma registrada.")

    # ---------------------------- ENTRADAS ----------------------------
    def register_entrada(self):
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name,unit_type FROM products WHERE active=1 ORDER BY name ASC")
            rows = c.fetchall()
            values = [f"{name}" for pid, name, ut in rows]
        dlg = EntradaDialog(self, values)
        self.wait_window(dlg)
        if not dlg.result:
            return
        res = self.resolve_product(dlg.result["product_text"])
        if not res:
            messagebox.showerror("Entrada", "Producto inválido.")
            return
        pid, name, price, exp, unit_type, stock_db = res
        qty_txt = dlg.result["qty_text"]
        date_txt = dlg.result["date_text"]
        note_txt = dlg.result["note"]
        qty = int(round(parse_decimal_kg(qty_txt) * 1000)) if (unit_type or "unit") == "kg" else self._to_int(qty_txt, 0)
        if qty <= 0:
            messagebox.showerror("Entrada", "Cantidad inválida.")
            return
        d_iso = parse_any_date_to_iso(date_txt) or date.today().strftime("%Y-%m-%d")
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("UPDATE products SET stock=stock+? WHERE id=?", (qty, pid))
            c.execute("INSERT INTO movimientos(product_id,qty,kind,note,date) VALUES (?,?,?,?,?)",
                      (pid, qty, "entrada", (note_txt or ""), d_iso))
        self.refresh_products()
        messagebox.showinfo("Entrada", f"Entrada registrada para '{name}'.")

    # ---------------------------- Productos CRUD ----------------------------
    def refresh_products(self):
        self.tree_prod.delete(*self.tree_prod.get_children())
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id,name,category,stock,price,expiry,barcode,unit_type FROM products WHERE active=1 ORDER BY id DESC")
            rows = c.fetchall()
        today = date.today()
        near_to = today + timedelta(days=NEAR_EXP_DAYS)
        for pid, name, cat, stock, price, exp, bar, utype in rows:
            tag = ()
            exp_show = "-"
            if exp:
                try:
                    d = datetime.strptime(exp, "%Y-%m-%d").date()
                    exp_show = d.strftime("%d-%m-%Y")
                    if d < today:
                        tag = ("expired",)
                    elif today <= d <= near_to:
                        tag = ("near",)
                except Exception:
                    exp_show = exp
            if not tag:
                s = int(stock or 0)
                if (utype or "unit") == "kg":
                    if s <= self.KG_ALERT_GRAMS:
                        tag = ("low",)
                else:
                    if s <= self.STOCK_RED:
                        tag = ("low",)
                    elif self.STOCK_YELLOW_MIN <= s <= self.STOCK_YELLOW_MAX:
                        tag = ("warn",)
            stock_cell = qty_label_kg(int(stock)) if (utype or "unit") == "kg" else str(int(stock))
            self.tree_prod.insert(
                "",
                "end",
                values=(pid, name, cat, stock_cell, peso_str(price), exp_show, bar or ""),
                tags=tag
            )
        stripe(self.tree_prod)

    def on_pick_product(self, _=None):
        sel = self.tree_prod.selection()
        if not sel:
            return
        pid, name, cat, stock, price, exp, bar = self.tree_prod.item(sel[0], "values")
        row = self.get_product(int(pid))
        utype = row[4] if row else "unit"

        self.p_name.delete(0, "end"); self.p_name.insert(0, name)
        self.p_cat.delete(0, "end"); self.p_cat.insert(0, cat)

        if (utype or "unit") == "kg":
            grams = int(row[5] or 0)
            self.p_stock.delete(0, "end")
            self.p_stock.insert(0, f"{grams / 1000.0:.3f}".rstrip("0").rstrip("."))
            self.p_type.set("Kilos (precio por kg)")
        else:
            self.p_stock.delete(0, "end")
            self.p_stock.insert(0, re.sub(r"[^\d\-]", "", str(stock)))
            self.p_type.set("Unidades")

        self.p_price.delete(0, "end")
        self.p_price.insert(0, re.sub(r"[^\d\-]", "", str(price)))

        self.p_exp.delete(0, "end")
        self.p_exp.insert(0, format_dd_mm_yyyy(row[3] if row else exp))

        self.p_bar.delete(0, "end")
        self.p_bar.insert(0, bar or "")

    def _validate_unit_type_change(self, pid, new_type):
        if pid is None:
            return True
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT unit_type FROM products WHERE id=?", (pid,))
            row = c.fetchone()
        if not row:
            return True
        old_type = (row[0] or "unit")
        if old_type != new_type:
            return messagebox.askyesno(
                "Tipo de unidad",
                f"El producto está como '{old_type}'. Vas a cambiar a '{new_type}'.\n\n"
                "Esto puede dejar el stock inconsistente.\n\n¿Continuar de todos modos?"
            )
        return True

    def _apply_price_rounding_entry(self):
        try:
            raw = self._to_int(self.p_price.get(), 0)
            adj = round_price_to_tens_rule_5(raw)
            if adj != raw:
                self.p_price.delete(0, "end")
                self.p_price.insert(0, str(adj))
        except Exception:
            pass

    def product_save(self):
        name = self.p_name.get().strip()
        if not name:
            messagebox.showerror("Producto", "Nombre obligatorio.")
            return
        cat = self.p_cat.get().strip() or None
        bar = self.p_bar.get().strip() or None
        exp_in = (self.p_exp.get() or "").strip()
        exp = parse_any_date_to_iso(exp_in) if exp_in else None
        if exp_in and not exp:
            messagebox.showerror("Fecha", "Fecha inválida. Prueba ej.: 25-12-2025 / 25/12/2025 / 2025-12-25.")
            return

        utype = "unit" if self.p_type.get().startswith("Uni") else "kg"
        stock = int(round(parse_decimal_kg(self.p_stock.get()) * 1000)) if utype == "kg" else self._to_int(self.p_stock.get(), 0)

        price_raw = self._to_int(self.p_price.get(), 0)
        if price_raw < 0:
            messagebox.showerror("Precio", "El precio no puede ser negativo.")
            return
        price = round_price_to_tens_rule_5(price_raw)

        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT id, unit_type FROM products WHERE lower(name)=?", (name.lower(),))
            row = c.fetchone()
            if row:
                pid = row[0]
                if not self._validate_unit_type_change(pid, utype):
                    return
                if messagebox.askyesno("Existente", f"'{name}' ya existe. ¿Sumar {self.p_stock.get()} al stock y actualizar datos?"):
                    with self._conn() as conn2:
                        c2 = conn2.cursor()
                        c2.execute(
                            """
                            UPDATE products SET stock=stock+?, category=?, price=?, expiry=?, barcode=?, unit_type=?, active=1 WHERE id=?
                            """,
                            (stock, cat, price, exp, bar, utype, pid)
                        )
                        conn2.commit()
                    self.refresh_products()
                    self._clear_pform()
                    return
                return

        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                INSERT INTO products(name,category,stock,price,expiry,barcode,unit_type,active)
                VALUES (?,?,?,?,?,?,?,1)
                """,
                (name, cat, stock, price, exp, bar, utype)
            )
            conn.commit()
        self.refresh_products()
        self._clear_pform()

    def product_update(self):
        sel = self.tree_prod.selection()
        if not sel:
            return
        pid = int(self.tree_prod.item(sel[0], "values")[0])

        name = self.p_name.get().strip()
        cat = self.p_cat.get().strip() or None
        bar = self.p_bar.get().strip() or None
        exp_in = (self.p_exp.get() or "").strip()
        exp = parse_any_date_to_iso(exp_in) if exp_in else None
        if exp_in and not exp:
            messagebox.showerror("Fecha", "Fecha inválida. Prueba ej.: 25-12-2025 / 25/12/2025 / 2025-12-25.")
            return

        utype = "unit" if self.p_type.get().startswith("Uni") else "kg"
        if not self._validate_unit_type_change(pid, utype):
            return

        stock = int(round(parse_decimal_kg(self.p_stock.get()) * 1000)) if utype == "kg" else self._to_int(self.p_stock.get(), 0)
        price_raw = self._to_int(self.p_price.get(), 0)
        if price_raw < 0:
            messagebox.showerror("Precio", "El precio no puede ser negativo.")
            return
        price = round_price_to_tens_rule_5(price_raw)

        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                UPDATE products SET name=?, category=?, stock=?, price=?, expiry=?, barcode=?, unit_type=?, active=1 WHERE id=?
                """,
                (name, cat, stock, price, exp, bar, utype, pid)
            )
            conn.commit()
        self.refresh_products()
        self._clear_pform()

    def product_delete(self):
        sel = self.tree_prod.selection()
        if not sel:
            return
        pid = self.tree_prod.item(sel[0], "values")[0]
        if not messagebox.askyesno("Eliminar", "¿Marcar producto como INACTIVO (soft-delete)?"):
            return
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("UPDATE products SET active=0 WHERE id=?", (pid,))
            conn.commit()
        self.refresh_products()

    def _clear_pform(self):
        for e in (self.p_name, self.p_cat, self.p_stock, self.p_price, self.p_exp, self.p_bar):
            e.delete(0, "end")
        self.p_type.set("Unidades")

    def product_filter(self):
        q = (self.p_filter.get() or "").strip()
        self.tree_prod.delete(*self.tree_prod.get_children())
        where, params = [], []
        m = re.match(r"^\s*id\s*:\s*(\d+)\s*$", q, re.I)
        if m:
            where.append("id=?")
            params.append(int(m.group(1)))
        else:
            m2 = re.match(r"^\s*(<=|>=|<|>|=)\s*(\d+)\s*$", q)
            if m2:
                where.append(f"stock {m2.group(1)} ?")
                params.append(int(m2.group(2)))
            else:
                exp_iso = parse_any_date_to_iso(q) if q else None
                if exp_iso:
                    where.append("ifnull(expiry,'')=?")
                    params.append(exp_iso)
                else:
                    if q.lower() in ("pronto", "vence", "caduca", "vencer", "vencimiento"):
                        d1 = date.today().strftime("%Y-%m-%d")
                        d2 = (date.today() + timedelta(days=NEAR_EXP_DAYS)).strftime("%Y-%m-%d")
                        where.append("(expiry IS NOT NULL AND expiry<>'' AND date(expiry) BETWEEN ? AND ?)")
                        params.extend([d1, d2])
                    elif q:
                        like = f"%{q.lower()}%"
                        where.append("(lower(name) LIKE ? OR lower(category) LIKE ? OR ifnull(barcode,'') LIKE ? OR CAST(id AS TEXT) LIKE ?)")
                        params.extend([like, like, like, like])

        sql = "SELECT id,name,category,stock,price,expiry,barcode,unit_type FROM products WHERE active=1"
        if where:
            sql += " AND " + " AND ".join(where)
        sql += " ORDER BY id DESC"
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(sql, params)
            rows = c.fetchall()
        today = date.today()
        near_to = today + timedelta(days=NEAR_EXP_DAYS)
        for pid, name, cat, stock, price, exp, bar, utype in rows:
            tag = ()
            exp_show = "-"
            if exp:
                try:
                    d = datetime.strptime(exp, "%Y-%m-%d").date()
                    exp_show = d.strftime("%d-%m-%Y")
                    if d < today:
                        tag = ("expired",)
                    elif today <= d <= near_to:
                        tag = ("near",)
                except Exception:
                    exp_show = exp
            if not tag:
                s = int(stock or 0)
                if (utype or "unit") == "kg":
                    if s <= self.KG_ALERT_GRAMS:
                        tag = ("low",)
                else:
                    if s <= self.STOCK_RED:
                        tag = ("low",)
                    elif self.STOCK_YELLOW_MIN <= s <= self.STOCK_YELLOW_MAX:
                        tag = ("warn",)
            stock_cell = qty_label_kg(int(stock)) if (utype or "unit") == "kg" else str(int(stock))
            self.tree_prod.insert(
                "",
                "end",
                values=(pid, name, cat, stock_cell, peso_str(price), exp_show, bar or ""),
                tags=tag
            )
        stripe(self.tree_prod)

    # ---------------------------- Historial / Excel (Ventas) ----------------------------
    def hist_today(self):
        t = date.today().strftime("%d-%m-%Y")
        self.h_from.delete(0, "end"); self.h_to.delete(0, "end")
        self.h_from.insert(0, t); self.h_to.insert(0, t)

    def hist_search(self):
        f = (self.h_from.get() or "").strip()
        t = (self.h_to.get() or "").strip()
        d1 = parse_any_date_to_iso(f)
        d2 = parse_any_date_to_iso(t)
        if not d1 or not d2:
            messagebox.showerror("Fechas", "Fechas inválidas. Ejemplos: 25-12-2025 / 25/12/2025 / 2025-12-25.")
            return
        self.tree_hist.delete(*self.tree_hist.get_children())
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                """
                SELECT id, date, payment_method, total
                FROM tickets
                WHERE date BETWEEN ? AND ?
                ORDER BY id ASC
                """,
                (d1, d2)
            )
            rows = c.fetchall()
        for tid, d, pm, total in rows:
            prod_txt = self.ticket_products_text(tid)
            self.tree_hist.insert("", "end",
                                  values=(tid, prod_txt, peso_str(total or 0), self.fmt_dd(d), pm or "efectivo"))
        stripe(self.tree_hist)

    def hist_export(self):
        f = (self.h_from.get() or "").strip()
        t = (self.h_to.get() or "").strip()
        d1 = parse_any_date_to_iso(f)
        d2 = parse_any_date_to_iso(t)
        if not d1 or not d2:
            messagebox.showerror("Fechas", "Fechas inválidas. Ejemplos: 25-12-2025 / 25/12/2025 / 2025-12-25.")
            return

        with self._conn() as conn:
            c = conn.cursor()
            q = """
            SELECT s.id, p.name, s.qty, s.price, p.unit_type, t.date AS d,
                   s.payment_method, s.discount, s.discount_note, s.ticket_id
            FROM sales s
            JOIN products p ON s.product_id = p.id
            JOIN tickets t ON s.ticket_id = t.id
            WHERE date(t.date) BETWEEN ? AND ?
            ORDER BY s.ticket_id ASC, s.id ASC
            """
            c.execute(q, (d1, d2))
            rows = c.fetchall()

        if not rows:
            messagebox.showinfo("Excel", "No hay ventas en el rango.")
            return

        title = f"Ventas desde {format_dd_mm_yyyy(d1)} hasta {format_dd_mm_yyyy(d2)}" if d1 != d2 \
            else f"Ventas del día {format_dd_mm_yyyy(d1)}"
        folder = SALES_DAILY_DIR if d1 == d2 else SALES_HIST_DIR
        os.makedirs(folder, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        ws.append([STORE_NAME])
        ws.append([title])
        ws.append([])
        ws.append(["Ticket", "ID", "Producto", "Cantidad", "Precio", "Fecha", "Método",
                   "Descuento", "Motivo descuento", "Total venta producto", "Productos (ticket)"])

        total_sum = 0
        for sale_id, name, qty, price, utype, d, pm, discount, dnote, ticket_id in rows:
            if (utype or "unit") == "kg":
                line = (int(qty) / 1000.0) * int(price)
                qty_txt = qty_label_kg(int(qty))
            else:
                line = int(qty) * int(price)
                qty_txt = qty_label_unit(int(qty))
            line_after = max(int(round(line)) - int(discount or 0), 0)
            total_sum += line_after
            prods_txt = self.ticket_products_text(int(ticket_id))
            ws.append([ticket_id, sale_id, name, qty_txt, "",
                       format_dd_mm_yyyy(d), (pm or "efectivo"), "",
                       (dnote or ""), "", prods_txt])
            r = ws.max_row
            ws.cell(row=r, column=5, value=int(price)).number_format = peso_fmt()
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=8, value=int(discount or 0)).number_format = peso_fmt()
            ws.cell(row=r, column=8).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=10, value=int(line_after)).number_format = peso_fmt()
            ws.cell(row=r, column=10).alignment = Alignment(horizontal="right")

        add_totals_block(ws, total_sum)
        style_header(ws, title)
        ws_autowidth(ws)

        fname = ("ventas_" + format_dd_mm_yyyy(d1).replace("-", "_") + ".xlsx") if folder == SALES_DAILY_DIR \
            else ("ventas_" + format_dd_mm_yyyy(d1).replace("-", "_") + "_a_" + format_dd_mm_yyyy(d2).replace("-", "_") + ".xlsx")
        path = os.path.join(folder, fname)
        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Excel", f"No se pudo guardar:\n{e}")
            return
        messagebox.showinfo("Excel", f"Archivo generado:\n{path}")

    # ---------------------------- Exportar Movimientos (Entradas/Ventas/Mermas) ----------------------------
    def mov_export_using_sales_range(self):
        f = (self.h_from.get() or "").strip()
        t = (self.h_to.get() or "").strip()
        d1 = parse_any_date_to_iso(f)
        d2 = parse_any_date_to_iso(t)
        if not d1 or not d2:
            messagebox.showerror("Fechas", "Fechas inválidas. Ejemplos: 25-12-2025 / 25/12/2025 / 2025-12-25.")
            return

        # Trae movimientos del período
        with self._conn() as conn:
            c = conn.cursor()
            base = """SELECT m.id, p.id, p.name, p.category, m.kind, m.qty, m.date, m.note, p.unit_type
                      FROM movimientos m JOIN products p ON m.product_id=p.id
                      WHERE date(m.date) BETWEEN ? AND ?
                      ORDER BY m.id ASC"""
            c.execute(base, (d1, d2))
            rows = c.fetchall()
            if not rows:
                messagebox.showinfo("Excel", "No hay movimientos en el rango.")
                return

        d1_dd = format_dd_mm_yyyy(d1)
        d2_dd = format_dd_mm_yyyy(d2)
        if d1 == d2:
            title = f"Movimientos (Entradas/Ventas/Mermas) del día {d1_dd}"
            folder = MOVS_DAILY_DIR
            fname = "movimientos_" + d1_dd.replace("-", "_") + ".xlsx"
        else:
            title = f"Movimientos (Entradas/Ventas/Mermas) desde {d1_dd} hasta {d2_dd}"
            folder = MOVS_HIST_DIR
            fname = "movimientos_" + d1_dd.replace("-", "_") + "_a_" + d2_dd.replace("-", "_") + ".xlsx"

        os.makedirs(folder, exist_ok=True)
        wb = openpyxl.Workbook()

        # Hoja Movimientos
        ws_mov = wb.active
        ws_mov.title = "Movimientos"
        ws_mov.append([STORE_NAME])
        ws_mov.append([title])
        ws_mov.append([])
        ws_mov.append(["ID", "Producto", "Categoría", "Tipo", "Cantidad", "Fecha", "Detalle"])

        # Resumen
        summary = {}
        totals_units = {"entrada": 0, "venta": 0, "merma": 0}
        totals_kg = {"entrada": 0.0, "venta": 0.0, "merma": 0.0}
        product_ids = set()

        # Orden: categoría -> nombre (categorías vacías al final)
        def _cat_key(cat):
            return (1, '') if not (cat or '').strip() else (0, (cat or '').lower())
        rows_sorted = sorted(rows, key=lambda r: (_cat_key(r[3]), (r[2] or '').lower(), r[0]))

        for mid, pid, pname, pcat, mkind, qty, d, note, utype in rows_sorted:
            product_ids.add(pid)

            # Cantidad
            if (utype or "unit") == "kg":
                qty_txt = qty_label_kg(int(qty))
                totals_kg[mkind] = totals_kg.get(mkind, 0.0) + (int(qty) / 1000.0)
            else:
                qty_txt = qty_label_unit(int(qty))
                totals_units[mkind] = totals_units.get(mkind, 0) + int(qty)

            # Ocultar "descuento" en Detalle para ventas (solo en Movimientos)
            det = note or ""
            if (mkind == "venta") and det and ("descuento" in det.lower()):
                det = ""

            ws_mov.append([mid, pname, pcat or "", mkind, qty_txt, format_dd_mm_yyyy(d), det])

            # Acumular resumen por producto
            s = summary.get(pid)
            if not s:
                s = summary[pid] = {
                    "name": pname, "category": pcat, "unit_type": (utype or "unit"),
                    "entrada_units": 0, "entrada_kg": 0.0,
                    "sold_units": 0, "sold_kg": 0.0,
                    "merma_units": 0, "merma_kg": 0.0,
                    "stock": None
                }
            if (utype or "unit") == "kg":
                if mkind == "entrada": s["entrada_kg"] += int(qty) / 1000.0
                elif mkind == "venta": s["sold_kg"] += int(qty) / 1000.0
                elif mkind == "merma": s["merma_kg"] += int(qty) / 1000.0
            else:
                if mkind == "entrada": s["entrada_units"] += int(qty)
                elif mkind == "venta": s["sold_units"] += int(qty)
                elif mkind == "merma": s["merma_units"] += int(qty)

        # Stock actual
        if product_ids:
            try:
                with self._conn() as conn:
                    c = conn.cursor()
                    qmarks = ",".join("?" for _ in product_ids)
                    c.execute(f"SELECT id, stock, unit_type FROM products WHERE id IN ({qmarks})", tuple(product_ids))
                    for pid2, stock_now, utype_now in c.fetchall():
                        if pid2 in summary:
                            summary[pid2]["stock"] = (int(stock_now or 0), (utype_now or "unit"))
            except Exception:
                pass

        # Hoja Resumen por producto (orden: categoría → nombre)
        ws_sum = wb.create_sheet("Resumen por producto")
        ws_sum.append([STORE_NAME]); ws_sum.append([title]); ws_sum.append([])
        ws_sum.append(["Producto", "Categoría", "Entradas (unid/kg)", "Ventas (unid/kg)",
                       "Mermas (unid/kg)", "Stock inicial", "Stock actual"])

        items = list(summary.items())
        def _sum_key(item):
            pid, s = item
            return _cat_key(s.get("category")), (s.get("name") or '').lower(), pid

        for pid, s in sorted(items, key=_sum_key):
            stock_actual = int(s["stock"][0]) if s.get("stock") else None
            if s["unit_type"] == "kg":
                stock_inicial_kg = None
                if stock_actual is not None:
                    stock_inicial_kg = (stock_actual / 1000.0) + s["sold_kg"] + s["merma_kg"] - s["entrada_kg"]
                entr_txt = f"{s['entrada_kg']:.3f} kg"
                sold_txt = f"{s['sold_kg']:.3f} kg"
                merma_txt = f"{s['merma_kg']:.3f} kg"
                stock_act_txt = qty_label_kg(stock_actual) if stock_actual is not None else "-"
                stock_ini_txt = f"{stock_inicial_kg:.3f} kg" if stock_inicial_kg is not None else "-"
            else:
                stock_inicial_units = None
                if stock_actual is not None:
                    stock_inicial_units = stock_actual + s["sold_units"] + s["merma_units"] - s["entrada_units"]
                entr_txt = f"{s['entrada_units']} unidades"
                sold_txt = f"{s['sold_units']} unidades"
                merma_txt = f"{s['merma_units']} unidades"
                stock_act_txt = str(stock_actual) if stock_actual is not None else "-"
                stock_ini_txt = str(stock_inicial_units) if stock_inicial_units is not None else "-"
            ws_sum.append([s["name"], s["category"] or "", entr_txt, sold_txt, merma_txt, stock_ini_txt, stock_act_txt])

        # Hoja Totales del período
        ws_tot = wb.create_sheet("Totales del período")
        ws_tot.append([STORE_NAME]); ws_tot.append([title]); ws_tot.append([])
        ws_tot.append(["Tipo", "Unidades", "Kilos"])
        style_section_row(ws_tot, ws_tot.max_row, total_cols=7)
        ws_tot.append(["Entradas", totals_units.get("entrada", 0), f"{totals_kg.get('entrada', 0.0):.3f} kg"])
        ws_tot.append(["Ventas", totals_units.get("venta", 0), f"{totals_kg.get('venta', 0.0):.3f} kg"])
        ws_tot.append(["Mermas", totals_units.get("merma", 0), f"{totals_kg.get('merma', 0.0):.3f} kg"])

        # Hoja Ranking Top 15
        ws_rank = wb.create_sheet("Ranking Top 15")
        ws_rank.append([STORE_NAME]); ws_rank.append([title]); ws_rank.append([])

        by_units = [s for s in summary.values() if s["sold_units"] > 0]
        by_units.sort(key=lambda x: (-x["sold_units"], (x.get("category") or '').lower(), (x.get("name") or '').lower()))

        by_kg = [s for s in summary.values() if s["sold_kg"] > 0]
        by_kg.sort(key=lambda x: (-x["sold_kg"], (x.get("category") or '').lower(), (x.get("name") or '').lower()))

        ws_rank.append(["Top 15 por unidades", "", "", ""])
        style_section_row(ws_rank, ws_rank.max_row, total_cols=4)
        ws_rank.append(["Posición", "Producto", "Categoría", "Unidades vendidas"])
        for i, s in enumerate(by_units[:15], start=1):
            ws_rank.append([i, s["name"], s["category"] or "", int(s["sold_units"])])
        ws_rank.append([])

        ws_rank.append(["Top 15 por kilos", "", "", ""])
        style_section_row(ws_rank, ws_rank.max_row, total_cols=4)
        ws_rank.append(["Posición", "Producto", "Categoría", "Kilos vendidos"])
        for i, s in enumerate(by_kg[:15], start=1):
            ws_rank.append([i, s["name"], s["category"] or "", f"{s['sold_kg']:.3f} kg"])

        # Estilos
        style_header(ws_mov, title); ws_autowidth(ws_mov)
        style_header(ws_sum, title); ws_autowidth(ws_sum)
        style_header(ws_tot, title); ws_autowidth(ws_tot)
        style_header(ws_rank, title); ws_autowidth(ws_rank)

        path = os.path.join(folder, fname)
        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("Excel", f"No se pudo guardar:\n{e}")
            return
        messagebox.showinfo("Excel", f"Archivo generado:\n{path}")

    # ---------------------------- Auto export (ayer) ----------------------------
    def auto_export_yesterday(self):
        y = date.today() - timedelta(days=1)
        y_iso = y.strftime("%Y-%m-%d")
        y_dd = y.strftime("%d-%m-%Y")
        with self._conn() as conn:
            c = conn.cursor()
            q = """
            SELECT s.id, p.name, s.qty, s.price, p.unit_type, t.date AS d,
                   s.payment_method, s.discount, s.discount_note, s.ticket_id
            FROM sales s
            JOIN products p ON s.product_id=p.id
            JOIN tickets t ON s.ticket_id=t.id
            WHERE t.date = ?
            ORDER BY s.ticket_id ASC, s.id ASC
            """
            c.execute(q, (y_iso,))
            rows = c.fetchall()
            if not rows:
                return

        path = os.path.join(SALES_DAILY_DIR, "ventas_" + y_dd.replace("-", "_") + ".xlsx")
        if os.path.exists(path):
            return

        title = f"Ventas del día {y_dd}"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        ws.append([STORE_NAME]); ws.append([title]); ws.append([])
        ws.append(["Ticket", "ID", "Producto", "Cantidad", "Precio", "Fecha", "Método",
                   "Descuento", "Motivo descuento", "Total venta producto", "Productos (ticket)"])

        total = 0
        for sale_id, name, qty, price, utype, d, pm, discount, dnote, ticket_id in rows:
            if (utype or "unit") == "kg":
                line = (int(qty) / 1000.0) * int(price)
                qty_txt = qty_label_kg(int(qty))
            else:
                line = int(qty) * int(price)
                qty_txt = qty_label_unit(int(qty))
            line_after = max(int(round(line)) - int(discount or 0), 0)
            total += line_after
            prods_txt = self.ticket_products_text(int(ticket_id))
            ws.append([ticket_id, sale_id, name, qty_txt, "",
                       format_dd_mm_yyyy(d), (pm or "efectivo"), "",
                       (dnote or ""), "", prods_txt])
            r = ws.max_row
            ws.cell(row=r, column=5, value=int(price)).number_format = peso_fmt()
            ws.cell(row=r, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=8, value=int(discount or 0)).number_format = peso_fmt()
            ws.cell(row=r, column=8).alignment = Alignment(horizontal="right")
            ws.cell(row=r, column=10, value=int(line_after)).number_format = peso_fmt()
            ws.cell(row=r, column=10).alignment = Alignment(horizontal="right")

        add_totals_block(ws, total)
        style_header(ws, title)
        ws_autowidth(ws)
        try:
            os.makedirs(SALES_DAILY_DIR, exist_ok=True)
            wb.save(path)
        except Exception:
            pass

    # ---------------------------- Ventana “Ventas del día (ventana)” ----------------------------
    def open_day_sales_window(self):
        if self.day_win and self.day_win.winfo_exists():
            try:
                self.day_win.deiconify()
                self.day_win.lift()
                self.day_win.focus_force()
                self.day_win.attributes("-topmost", True)
                self.day_win.after(200, lambda: self.day_win.attributes("-topmost", False))
            except Exception:
                pass
            return

        win = ctk.CTkToplevel(self)
        win.title("Ventas del día")
        win.geometry("940x620")
        self.day_win = win
        try:
            win.transient(self)
            win.lift()
            win.focus_force()
            win.attributes("-topmost", True)
            win.after(200, lambda: win.attributes("-topmost", False))
        except Exception:
            pass

        def _on_close():
            if self.day_win_refresh_job:
                try:
                    self.after_cancel(self.day_win_refresh_job)
                except Exception:
                    pass
            self.day_win_refresh_job = None
            self.day_win_controls = None
            self.day_win = None
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", _on_close)

        frame = ctk.CTkFrame(win, fg_color="#FFFFFF")
        frame.pack(fill="both", expand=True, padx=8, pady=8)

        filt = ctk.CTkFrame(frame, fg_color="#FFFFFF")
        filt.pack(fill="x", padx=4, pady=(0, 6))
        ctk.CTkLabel(filt, text="Buscar por:", text_color="#111827").pack(side="left", padx=(6, 4))
        cb = ctk.CTkComboBox(filt, values=["hora", "fecha", "producto"], width=160)
        cb.set("fecha")
        cb.pack(side="left", padx=4)
        ent = ctk.CTkEntry(filt, width=220)
        ent.pack(side="left", padx=6)
        ctk.CTkButton(filt, text="Buscar", fg_color="#2C6FB7",
                      command=lambda: self._filter_day_window(tree, cb.get(), ent.get())).pack(side="left", padx=6)
        ctk.CTkButton(filt, text="Hoy", command=lambda: ent.delete(0, "end")).pack(side="left", padx=6)

        cols = ("Ticket", "Productos", "Total", "Fecha", "Hora", "Método", "Detalle")
        day_box, tree = wrap_treeview(frame)
        day_box.pack(fill="both", expand=True)
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c, anchor="center")
            tree.column(c, width=130, anchor="center")
        tree.column("Productos", width=380, anchor="w")
        tree.column("Hora", width=100, anchor="center")
        tree.column("Método", width=120, anchor="center")
        tree.column("Detalle", width=120, anchor="center")

        # Eventos CORREGIDOS
        tree.bind("<Double-Button-1>", self._on_tree_window_click)
        tree.bind("<ButtonPress-1>", lambda e: setattr(self, "day_win_paused", True))
        tree.bind("<ButtonRelease-1>", lambda e: self.after(400, lambda: setattr(self, "day_win_paused", False)))

        tot_bar = ctk.CTkFrame(frame, fg_color="#FFFFFF")
        tot_bar.pack(fill="x", pady=(6, 0))
        self.win_neto = ctk.CTkLabel(tot_bar, text="Neto: $ 0", font=("Segoe UI", 12, "bold"))
        self.win_iva = ctk.CTkLabel(tot_bar, text="IVA: $ 0", font=("Segoe UI", 12, "bold"))
        self.win_total = ctk.CTkLabel(tot_bar, text="Total: $ 0", font=("Segoe UI", 14, "bold"), text_color="#2F80ED")
        self.win_neto.pack(side="left", padx=6)
        self.win_iva.pack(side="left", padx=12)
        self.win_total.pack(side="left", padx=14)

        ctk.CTkButton(frame, text="Eliminar venta", fg_color="#C91414",
                      command=lambda: self._delete_from_window(tree)).pack(side="right", padx=6, pady=6)

        self.day_win_controls = {"tree": tree, "cb": cb, "ent": ent}
        self._filter_day_window(tree, "fecha", date.today().strftime("%d-%m-%Y"))

        def _auto_refresh():
            if not (self.day_win and self.day_win.winfo_exists()):
                return
            if getattr(self, "day_win_paused", False):
                self.day_win_refresh_job = self.after(2000, _auto_refresh)
                return
            try:
                tree2 = self.day_win_controls["tree"]
                mode = self.day_win_controls["cb"].get()
                txt = self.day_win_controls["ent"].get()
                self._filter_day_window(tree2, mode, txt)
            except Exception:
                pass
            finally:
                self.day_win_refresh_job = self.after(2000, _auto_refresh)
        _auto_refresh()

    def refresh_day_window_if_open(self):
        if self.day_win and self.day_win.winfo_exists() and self.day_win_controls:
            tree = self.day_win_controls["tree"]
            mode = self.day_win_controls["cb"].get()
            txt = self.day_win_controls["ent"].get()
            self._filter_day_window(tree, mode, txt)

    def _delete_from_window(self, tree):
        sel = tree.selection()
        if not sel:
            return
        tid = tree.item(sel[0], "values")[0]
        if not messagebox.askyesno("Eliminar", "¿Eliminar este TICKET y restaurar stock?"):
            return
        with self._conn() as conn:
            c = conn.cursor()
            c.execute("SELECT product_id, qty FROM sales WHERE ticket_id=?", (tid,))
            for pid, q in c.fetchall():
                c.execute("UPDATE products SET stock=stock+? WHERE id=?", (int(q), int(pid)))
            c.execute("DELETE FROM sales WHERE ticket_id=?", (tid,))
            c.execute("DELETE FROM tickets WHERE id=?", (tid,))
            conn.commit()
        tree.delete(sel[0])

    def _filter_day_window(self, tree, mode, text):
        selected_ids = []
        for iid in tree.selection():
            try:
                val_tid = tree.item(iid, "values")[0]
                selected_ids.append(int(val_tid))
            except Exception:
                pass

        tree.delete(*tree.get_children())
        d_iso = date.today().strftime("%Y-%m-%d")
        if mode == "fecha" and text:
            iso_try = parse_any_date_to_iso(text.strip())
            if not iso_try:
                messagebox.showerror("Fecha", "Usa formato válido (25-12-2025, 25/12/2025, 2025-12-25).")
                return
            d_iso = iso_try

        with self._conn() as conn:
            c = conn.cursor()
            c.execute("""SELECT id, date, created_at, payment_method, total FROM tickets WHERE date=? ORDER BY id ASC""", (d_iso,))
            trows = c.fetchall()

        keep_ids = None
        if mode == "producto" and (text or "").strip():
            q = text.strip().lower()
            with self._conn() as conn:
                c = conn.cursor()
                c.execute(
                    """
                    SELECT DISTINCT t.id
                    FROM tickets t
                    JOIN sales s ON s.ticket_id = t.id
                    JOIN products p ON p.id = s.product_id
                    WHERE t.date=? AND lower(p.name) LIKE ?
                    """,
                    (d_iso, f"%{q}%")
                )
                keep_ids = set(r[0] for r in c.fetchall())

        subtotal = 0
        for tid, d, created_at, pm, total in trows:
            if keep_ids is not None and tid not in keep_ids:
                continue
            hora_txt = "--"
            if created_at:
                try:
                    hora_txt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S").strftime("%H:%M")
                except Exception:
                    hora_txt = created_at.split(" ")[1] if " " in created_at else created_at
            prod_txt = self.ticket_products_text(tid)
            subtotal += int(total or 0)
            tree.insert("", "end", iid=str(tid),
                        values=(tid, prod_txt, peso_str(total or 0), format_dd_mm_yyyy(d), hora_txt, pm or "efectivo", "Ver detalle"))
        stripe(tree)
        neto = int(round(subtotal / (1 + IVA)))
        iva = subtotal - neto
        self.win_neto.configure(text=f"Neto: {peso_str(neto)}")
        self.win_iva.configure(text=f"IVA: {peso_str(iva)}")
        self.win_total.configure(text=f"Total: {peso_str(subtotal)}")

        for tid in selected_ids:
            sid = str(tid)
            try:
                tree.selection_add(sid)
                tree.see(sid)
            except Exception:
                pass

    def _on_tree_window_click(self, event):
        widget = event.widget
        if not isinstance(widget, ttk.Treeview):
            return
        region = widget.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = widget.identify_column(event.x)
        row = widget.identify_row(event.y)
        if not row:
            return
        if col == "#7":
            values = widget.item(row, "values")
            if values and len(values) >= 1:
                try:
                    self.day_win_paused = True
                    ticket_id = int(values[0])
                    self.show_ticket_detail(ticket_id)
                    self.after(1000, lambda: setattr(self, "day_win_paused", False))
                except Exception:
                    pass

    # ---------------------------- Sugerencias Producto ----------------------------
    def _on_product_typed_autocomplete(self, event=None):
        if event and event.keysym in ("Up", "Down", "Left", "Right", "Escape", "Return", "Tab"):
            return
        txt = (self.ent_prod.get() or "").strip().lower()
        if not txt:
            self._suggest_popup.hide()
            return
        with self._conn() as conn:
            c = conn.cursor()
            c.execute(
                "SELECT id,name,price FROM products WHERE active=1 AND lower(name) LIKE ? ORDER BY name ASC LIMIT 50",
                (f"%{txt}%",)
            )
            rows = c.fetchall()
        suggestions = [f"{name} — {peso_str(price)}" for pid, name, price in rows]
        self._suggest_popup.show(suggestions)

    def _on_suggestion_selected(self, text: str, add_to_cart: bool = False):
        name_only = text.split(" — ")[0].strip() if " — " in text else text.strip()
        self.ent_prod.delete(0, "end")
        self.ent_prod.insert(0, name_only)
        if add_to_cart:
            self.add_to_cart()

# #### - - - - RUN - - - - 
if __name__ == "__main__":
    try:
        app = App()
        app.mainloop()
    except Exception as e:
        import traceback
        error_msg = f"Error no capturado: {str(e)}\n\n{traceback.format_exc()}"
        print(error_msg)
        with open(os.path.join(BASE_DIR(), "error_log.txt"), "a", encoding="utf-8") as f:
            f.write(f"\n{'='*60}\n{datetime.now()}\n{error_msg}\n")
        try:
            messagebox.showerror("Error", error_msg)
        except:
            pass
